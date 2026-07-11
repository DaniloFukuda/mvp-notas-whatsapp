import sys
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import web_upload
from services.rdv_service import RDVService


EXPECTED_SHEETS = (
    "Lancamentos",
    "Resumo por Colaborador",
    "Resumo por Categoria",
    "Pendencias",
)


def main() -> None:
    original_service = web_upload.rdv_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            service = RDVService(Path(temp_dir) / "rdv_excel_test.db")
            web_upload.rdv_service = service
            danilo = service.get_collaborator_by_phone("5500000000001")
            marcelo = service.get_collaborator_by_phone("5500000000002")
            henrique = service.get_collaborator_by_phone("5500000000003")
            assert danilo is not None
            assert marcelo is not None
            assert henrique is not None

            first = service.create_whatsapp_receipt(
                collaborator_id=danilo["id"],
                phone=danilo["telefone_whatsapp"],
                input_type="imagem",
                file_path="data/documentos/uploads/whatsapp/demo_combustivel.jpg",
                whatsapp_message_id="wamid.excel.complete",
                received_at="2026-06-09T08:30:00",
                observation="visita tecnica",
                analysis={
                    "valor_detectado": 215.40,
                    "data_detectada": "2026-06-09",
                    "fornecedor_detectado": "POSTO FICTICIO LTDA",
                    "chave_acesso": "1" * 44,
                    "origem_valor": "ocr",
                },
            )
            first = service.complete_launch_category(first["id"], "combustivel")
            assert first["status_revisao"] == "aprovado"
            assert first["falha_leitura"] == 0

            manual = service.register_manual_expense(
                colaborador="Danilo",
                data_despesa="2026-06-11",
                categoria="manutencao",
                valor="50,00",
                cidade_origem="Ribeirao Preto",
                cidade_destino="Sertaozinho",
                km_inicio="1000",
                km_fim="1120",
                observacao="deslocamento de teste",
            )
            service.update_review_status(manual["id"], "aprovado")

            second = service.create_whatsapp_receipt(
                collaborator_id=marcelo["id"],
                phone=marcelo["telefone_whatsapp"],
                input_type="documento",
                file_path="data/documentos/uploads/whatsapp/demo_hotel.pdf",
                whatsapp_message_id="wamid.excel.review",
                received_at="2026-06-10T18:15:00",
                observation="aguardando conferencia",
            )
            service.save_launch_value(second["id"], "380,00")
            service.complete_launch_category(second["id"], "hospedagem")
            service.mark_launch_for_review(second["id"])

            open_trip = service.create_whatsapp_km_launch(
                collaborator_id=henrique["id"],
                phone=henrique["telefone_whatsapp"],
                km_start="50000",
                received_at="2026-06-10T19:00:00",
            )
            assert open_trip["status_fluxo"] == "viagem_em_andamento"

            response = web_upload.exportar_relatorio_semanal_rdv_excel(
                semana="2026-W24",
                colaborador_id="",
                status="",
            )
            assert response.media_type == (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            assert response.headers["content-disposition"].endswith(
                'rdv_ciclus_agro_2026_W24.xlsx"'
            )
            assert response.body.startswith(b"PK")
            assert any(
                route.path == "/ciclus/rdv/relatorio-semanal.xlsx"
                for route in web_upload.app.routes
            )
            page = web_upload.listar_rdv_ciclus(
                colaborador_id="",
                status="",
                semana="2026-W24",
            )
            assert "Baixar relatorio Excel" in page
            assert (
                '/ciclus/rdv/relatorio-semanal.xlsx?semana=2026-W24'
                in page
            )

            workbook = load_workbook(BytesIO(response.body))
            assert tuple(workbook.sheetnames) == EXPECTED_SHEETS

            launches = workbook["Lancamentos"]
            assert _headers(launches) == (
                "Data",
                "Colaborador",
                "Telefone",
                "Categoria",
                "Valor",
                "Fornecedor detectado",
                "Data detectada",
                "Origem do valor",
                "Chave de acesso / QR Code / URL",
                "Cidade origem",
                "Cidade destino",
                "KM inicial",
                "KM final",
                "KM rodado",
                "Status",
                "Observacao",
                "Tipo de entrada",
                "Comprovante/arquivo",
                "Recebido em",
            )
            assert launches.max_row == 5
            assert launches.freeze_panes == "A2"
            assert launches.auto_filter.ref == launches.dimensions
            assert launches["E2"].number_format == 'R$ #,##0.00'
            assert launches["A2"].number_format == "dd/mm/yyyy"
            assert launches["S2"].number_format == "dd/mm/yyyy hh:mm"
            assert {
                launches["R2"].value,
                launches["R3"].value,
                launches["R4"].value,
                launches["R5"].value,
            } == {
                "demo_combustivel.jpg",
                "demo_hotel.pdf",
                None,
            }
            detected_row = next(
                row
                for row in range(2, launches.max_row + 1)
                if launches.cell(row, 18).value == "demo_combustivel.jpg"
            )
            assert launches.cell(detected_row, 6).value == "POSTO FICTICIO LTDA"
            assert launches.cell(detected_row, 7).number_format == "dd/mm/yyyy"
            assert launches.cell(detected_row, 8).value == "Ocr"
            assert launches.cell(detected_row, 9).value == "1" * 44
            km_row = next(
                row
                for row in range(2, launches.max_row + 1)
                if launches.cell(row, 10).value == "Ribeirao Preto"
            )
            assert launches.cell(km_row, 11).value == "Sertaozinho"
            assert launches.cell(km_row, 12).value == 1000
            assert launches.cell(km_row, 13).value == 1120
            assert launches.cell(km_row, 14).value == 120
            open_trip_row = next(
                row
                for row in range(2, launches.max_row + 1)
                if launches.cell(row, 12).value == 50000
            )
            assert launches.cell(open_trip_row, 10).value is None
            assert launches.cell(open_trip_row, 11).value is None
            assert launches.cell(open_trip_row, 12).value == 50000
            assert launches.cell(open_trip_row, 13).value is None
            assert launches.cell(open_trip_row, 14).value is None
            assert "Viagem Em Andamento" in launches.cell(open_trip_row, 15).value

            collaborators = workbook["Resumo por Colaborador"]
            assert _headers(collaborators) == (
                "Colaborador",
                "Total em R$",
                "Quantidade de lancamentos",
                "Quilometragem total",
                "Pendentes",
            )
            assert collaborators["B2"].number_format == 'R$ #,##0.00'
            collaborator_rows = {
                row[0].value: tuple(cell.value for cell in row[1:])
                for row in collaborators.iter_rows(min_row=2)
            }
            assert collaborator_rows["Danilo"] == (265.4, 2, 120, 0)
            assert collaborator_rows["Marcelo"] == (380, 1, 0, 1)
            assert collaborator_rows["Henrique Saraiva"] == (0, 1, 0, 1)

            categories = workbook["Resumo por Categoria"]
            assert _headers(categories) == (
                "Categoria",
                "Total em R$",
                "Quantidade",
            )
            assert categories.max_row == 5

            pending = workbook["Pendencias"]
            assert pending.max_row == 3
            assert {pending["B2"].value, pending["B3"].value} == {
                "Marcelo",
                "Henrique Saraiva",
            }
            assert "demo_hotel.pdf" in {pending["R2"].value, pending["R3"].value}
            assert all(
                cell.value != "demo_combustivel.jpg"
                for cell in pending["R"][1:]
            )
            assert pending.freeze_panes == "A2"
            assert pending.auto_filter.ref == pending.dimensions

            for sheet in workbook.worksheets:
                assert sheet.freeze_panes == "A2"
                assert sheet.auto_filter.ref == sheet.dimensions
                assert all(
                    sheet.column_dimensions[column].width > 0
                    for column in sheet.column_dimensions
                )
                assert all(cell.font.bold for cell in sheet[1])
    finally:
        web_upload.rdv_service = original_service

    print("OK: exportacao Excel semanal do RDV validada.")


def _headers(sheet) -> tuple:
    return tuple(cell.value for cell in sheet[1])


if __name__ == "__main__":
    main()
