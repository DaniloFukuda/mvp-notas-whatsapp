import sys
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from services.visitas_excel_service import build_visitas_workbook
from services.visitas_service import VisitasTecnicasService


def test_visita_planilha_exporta_abas_e_link_gps():
    with tempfile.TemporaryDirectory() as temp_dir:
        service = VisitasTecnicasService(Path(temp_dir) / "visitas.db")
        visita = service.iniciar_visita("5500000000001", tecnico_nome="Danilo")
        service.atualizar_campo(visita["id"], "data_visita", "2026-06-17")
        service.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
        service.atualizar_campo(visita["id"], "area_hectares", 2299)
        service.adicionar_localizacao(visita["id"], -15.0019124, -50.7714295)
        service.adicionar_midia(
            visita["id"],
            "foto",
            media_id_whatsapp="wamid.foto",
            caminho_arquivo="data/documentos/uploads/whatsapp/foto.jpg",
            legenda="Tanque",
        )
        service.adicionar_dado_coletado(
            visita["id"],
            "tanque",
            "capacidade 10000 L",
        )

        content = build_visitas_workbook(service.listar_visitas(mes="2026-06"))
        workbook = load_workbook(BytesIO(content))

        assert tuple(workbook.sheetnames) == (
            "Visitas",
            "Fotos",
            "Localizações",
            "Dados coletados",
        )
        visitas = workbook["Visitas"]
        assert visitas.cell(1, 15).value == "Link GPS principal"
        assert visitas.cell(1, 3).value == "Técnico"
        assert visitas.cell(1, 6).value == "Proprietário"
        assert visitas.cell(1, 8).value == "Área ha"
        assert visitas.cell(2, 5).value == "Fazenda Imperial"
        assert visitas.cell(2, 13).value == 1
        assert visitas.cell(2, 14).value == 1
        assert visitas.cell(2, 15).value == (
            "https://maps.google.com/?q=-15.0019124,-50.7714295"
        )
        assert workbook["Fotos"].cell(2, 9).value == "foto.jpg"
        assert workbook["Localizações"].cell(1, 4).value == "Descrição"
        assert workbook["Localizações"].cell(2, 7).value.startswith(
            "https://maps.google.com/?q="
        )
        assert workbook["Dados coletados"].cell(2, 4).value == "tanque"


def test_planilha_visitas_exclui_canceladas():
    with tempfile.TemporaryDirectory() as temp_dir:
        service = VisitasTecnicasService(Path(temp_dir) / "visitas.db")
        cancelada = service.iniciar_visita("5500000000001", tecnico_nome="Danilo")
        service.atualizar_campo(cancelada["id"], "data_visita", "2026-06-17")
        service.atualizar_campo(cancelada["id"], "fazenda", "Fazenda Cancelada")
        service.cancelar_visita(cancelada["id"])
        fechada = service.iniciar_visita("5500000000001", tecnico_nome="Danilo")
        service.atualizar_campo(fechada["id"], "data_visita", "2026-06-17")
        service.atualizar_campo(fechada["id"], "fazenda", "Fazenda Valida")
        service.fechar_visita(fechada["id"])

        content = build_visitas_workbook(service.listar_visitas(mes="2026-06"))
        workbook = load_workbook(BytesIO(content))
        visitas = workbook["Visitas"]

        fazendas = [
            row[0]
            for row in visitas.iter_rows(
                min_row=2,
                min_col=5,
                max_col=5,
                values_only=True,
            )
        ]
        assert "Fazenda Valida" in fazendas
        assert "Fazenda Cancelada" not in fazendas


def test_planilha_visitas_global():
    with tempfile.TemporaryDirectory() as temp_dir:
        service = VisitasTecnicasService(Path(temp_dir) / "visitas.db")
        primeira = service.iniciar_visita("5500000000001", tecnico_nome="Danilo")
        service.atualizar_campo(primeira["id"], "data_visita", "2026-05-10")
        service.atualizar_campo(primeira["id"], "fazenda", "Fazenda Imperial")
        segunda = service.iniciar_visita("5500000000002", tecnico_nome="Marcelo")
        service.atualizar_campo(segunda["id"], "data_visita", "2026-06-17")
        service.atualizar_campo(segunda["id"], "fazenda", "Fazenda Boi Dourado 3J")

        content = build_visitas_workbook(service.listar_visitas_validas())
        workbook = load_workbook(BytesIO(content))
        visitas = workbook["Visitas"]

        fazendas = [
            row[0]
            for row in visitas.iter_rows(
                min_row=2,
                min_col=5,
                max_col=5,
                values_only=True,
            )
        ]
        assert "Fazenda Imperial" in fazendas
        assert "Fazenda Boi Dourado 3J" in fazendas
