import sys
import tempfile
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import web_upload
from services.rdv_service import RDVService


EXPECTED_SHEETS = (
    "Lancamentos",
    "Resumo por Colaborador",
    "Resumo por Categoria",
    "Pendencias",
)


def test_weekly_rdv_excel_export_contains_expenses_km_and_pending_rows():
    original_service = web_upload.rdv_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            service = RDVService(Path(temp_dir) / "rdv_excel_test.db")
            web_upload.rdv_service = service
            danilo = service.get_collaborator_by_phone("5500000000001")
            marcelo = service.get_collaborator_by_phone("5500000000002")
            henrique = service.get_collaborator_by_phone("5500000000003")

            first = service.create_whatsapp_receipt(
                collaborator_id=danilo["id"],
                phone=danilo["telefone_whatsapp"],
                input_type="imagem",
                file_path="data/documentos/uploads/whatsapp/demo_combustivel.jpg",
                whatsapp_message_id="wamid.excel.complete",
                received_at="2026-06-09T08:30:00",
                observation="visita tecnica",
                analysis={
                    "valor_detectado": 215.40,
                    "data_detectada": "2026-06-09",
                    "fornecedor_detectado": "POSTO FICTICIO LTDA",
                    "chave_acesso": "1" * 44,
                    "origem_valor": "ocr",
                },
            )
            first = service.complete_launch_category(first["id"], "combustivel")
            assert first["status_revisao"] == "aprovado"

            manual = service.register_manual_expense(
                colaborador="Danilo",
                data_despesa="2026-06-11",
                categoria="manutencao",
                valor="50,00",
                cidade_origem="Ribeirao Preto",
                cidade_destino="Sertaozinho",
                km_inicio="1000",
                km_fim="1120",
                observacao="deslocamento de teste",
            )
            service.update_review_status(manual["id"], "aprovado")

            second = service.create_whatsapp_receipt(
                collaborator_id=marcelo["id"],
                phone=marcelo["telefone_whatsapp"],
                input_type="documento",
                file_path="data/documentos/uploads/whatsapp/demo_hotel.pdf",
                whatsapp_message_id="wamid.excel.review",
                received_at="2026-06-10T18:15:00",
                observation="aguardando conferencia",
            )
            service.save_launch_value(second["id"], "380,00")
            service.save_launch_receipt_date(second["id"], "10/06/2026")
            service.complete_launch_category(second["id"], "hospedagem")
            service.mark_launch_for_review(second["id"])

            open_trip = service.create_whatsapp_km_launch(
                collaborator_id=henrique["id"],
                phone=henrique["telefone_whatsapp"],
                km_start="50000",
                received_at="2026-06-10T19:00:00",
            )
            service.save_km_origin(open_trip["id"], "Formosa")
            open_trip = service.save_km_destination(
                open_trip["id"], "Fazenda Santa Rita"
            )
            assert open_trip["status_fluxo"] == "viagem_em_andamento"

            response = web_upload.exportar_relatorio_semanal_rdv_excel(
                semana="2026-W24",
                colaborador_id="",
                status="",
            )
            assert response.media_type == (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            assert response.body.startswith(b"PK")

            workbook = load_workbook(BytesIO(response.body))
            assert tuple(workbook.sheetnames) == EXPECTED_SHEETS

            launches = workbook["Lancamentos"]
            assert _headers(launches)[0:6] == (
                "Data do comprovante",
                "Data de envio",
                "Colaborador",
                "Telefone",
                "Categoria",
                "Valor",
            )
            assert launches.max_row == 5
            detected_row = next(
                row
                for row in range(2, launches.max_row + 1)
                if launches.cell(row, 19).value == "demo_combustivel.jpg"
            )
            assert launches.cell(detected_row, 7).value == "POSTO FICTICIO LTDA"
            assert _date_value(launches.cell(detected_row, 1).value) == date(
                2026, 6, 9
            )
            assert launches.cell(detected_row, 2).value == datetime(2026, 6, 9, 8, 30)
            assert _date_value(launches.cell(detected_row, 8).value) == date(
                2026, 6, 9
            )
            assert launches.cell(detected_row, 8).number_format == "dd/mm/yyyy"

            km_row = next(
                row
                for row in range(2, launches.max_row + 1)
                if launches.cell(row, 11).value == "Ribeirao Preto"
            )
            assert launches.cell(km_row, 12).value == "Sertaozinho"
            assert launches.cell(km_row, 15).value == 120

            manual_date_row = next(
                row
                for row in range(2, launches.max_row + 1)
                if launches.cell(row, 19).value == "demo_hotel.pdf"
            )
            assert _date_value(launches.cell(manual_date_row, 1).value) == date(
                2026, 6, 10
            )
            assert launches.cell(manual_date_row, 2).value == datetime(
                2026, 6, 10, 18, 15
            )
            assert _date_value(launches.cell(manual_date_row, 8).value) == date(
                2026, 6, 10
            )

            open_trip_row = next(
                row
                for row in range(2, launches.max_row + 1)
                if launches.cell(row, 13).value == 50000
            )
            assert launches.cell(open_trip_row, 11).value == "Formosa"
            assert launches.cell(open_trip_row, 12).value == "Fazenda Santa Rita"
            assert "Viagem Em Andamento" in launches.cell(open_trip_row, 16).value
    finally:
        web_upload.rdv_service = original_service


def test_monthly_rdv_excel_export_uses_receipt_date_and_audit_sheet():
    original_service = web_upload.rdv_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            service = RDVService(Path(temp_dir) / "rdv_monthly_excel_test.db")
            web_upload.rdv_service = service
            danilo = service.get_collaborator_by_phone("5500000000001")

            receipt = service.create_whatsapp_receipt(
                collaborator_id=danilo["id"],
                phone=danilo["telefone_whatsapp"],
                input_type="imagem",
                file_path="data/documentos/uploads/whatsapp/mercado_pago.jpg",
                whatsapp_message_id="wamid.monthly.excel",
                received_at="2026-06-16T18:59:00",
                analysis={
                    "valor_detectado": 80,
                    "data_detectada": "2026-06-14",
                    "fornecedor_detectado": "Mercado Pago",
                    "origem_valor": "ocr",
                },
            )
            service.complete_launch_category(receipt["id"], "alimentacao")

            response = web_upload.exportar_relatorio_mensal_rdv_excel(
                mes="2026-06",
                colaborador_id="",
                status="",
            )

            assert response.media_type == (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            workbook = load_workbook(BytesIO(response.body))
            assert tuple(workbook.sheetnames) == (
                "Lancamentos",
                "Resumo por Colaborador",
                "Resumo por Categoria",
                "Pendencias",
                "Auditoria",
            )

            launches = workbook["Lancamentos"]
            headers = _headers(launches)
            assert headers[0] == "Data do comprovante"
            assert headers[1] == "Data de envio"
            assert "Data detectada" not in headers
            assert _date_value(launches.cell(2, 1).value) == date(2026, 6, 14)
            assert launches.cell(2, 2).value == datetime(2026, 6, 16, 18, 59)
            assert launches.cell(2, 6).value == 80
            assert launches.cell(2, 7).value == "Mercado Pago"

            audit = workbook["Auditoria"]
            assert _headers(audit) == (
                "id",
                "data_detectada",
                "recebido_em",
                "whatsapp_message_id",
                "caminho_arquivo",
            )
            assert _date_value(audit.cell(2, 2).value) == date(2026, 6, 14)
            assert audit.cell(2, 3).value == datetime(2026, 6, 16, 18, 59)
            assert audit.cell(2, 4).value == "wamid.monthly.excel"
    finally:
        web_upload.rdv_service = original_service


def _headers(sheet) -> tuple:
    return tuple(cell.value for cell in sheet[1])


def _date_value(value):
    return value.date() if isinstance(value, datetime) else value
