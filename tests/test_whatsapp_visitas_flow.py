import sys
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader
import pytest


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import api_whatsapp
from services.rdv_service import RDVService
from services.visitas_service import VisitasTecnicasService


def _install_services(temp_dir):
    rdv = RDVService(Path(temp_dir) / "rdv.db")
    visitas = VisitasTecnicasService(Path(temp_dir) / "visitas.db")
    api_whatsapp.rdv_service = rdv
    api_whatsapp.visitas_service = visitas
    api_whatsapp.visita_edit_states.clear()
    api_whatsapp.visita_active_states.clear()
    api_whatsapp.visita_new_visit_states.clear()
    collaborator = rdv.get_collaborator_by_phone("5500000000001")
    return rdv, visitas, collaborator["telefone_whatsapp"]


def _extract_pdf_text(content):
    reader = PdfReader(BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _create_test_visit(
    visitas,
    index,
    *,
    fazenda=None,
    data_visita=None,
    status="fechada",
    tecnico_nome=None,
    gerente=None,
):
    visita = visitas.criar_visita(
        f"5500000009{index:03d}",
        tecnico_nome=tecnico_nome or f"Tecnico {index}",
        fazenda=fazenda or f"Fazenda {index:02d}",
        estado_fluxo="visita_aberta",
    )
    visitas.atualizar_campo(visita["id"], "data_visita", data_visita or f"2026-07-{index:02d}")
    visitas.atualizar_campo(visita["id"], "gerente", gerente or f"Gerente {index}")
    if status == "fechada":
        visita = visitas.fechar_visita(visita["id"])
    elif status != "aberta":
        visitas.atualizar_campo(visita["id"], "status", status)
    return visitas.obter_visita(visita["id"])


def _ids_in_message_text(text):
    return [
        int(line.split(" ", 1)[0][1:])
        for line in text.splitlines()
        if line.startswith("#")
    ]


def test_visita_iniciar_fluxo():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            reply = api_whatsapp.handle_rdv_text_message(sender, "visita")

            assert "Vamos iniciar uma visita técnica." in reply
            assert "Qual o nome da fazenda?" in reply
            assert visitas.obter_visita_aberta(sender)["estado_fluxo"] == "aguardando_fazenda"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_preencher_campos():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            assert "proprietário" in api_whatsapp.handle_rdv_text_message(
                sender, "Fazenda Imperial"
            ).lower()
            api_whatsapp.handle_rdv_text_message(sender, "Alexander Duarte Paniago")
            api_whatsapp.handle_rdv_text_message(sender, "Paulo Silva")
            api_whatsapp.handle_rdv_text_message(sender, "2299 ha")
            api_whatsapp.handle_rdv_text_message(sender, "26/27")
            final = api_whatsapp.handle_rdv_text_message(
                sender, "Apresentacao de produtos ao cliente"
            )

            visita = visitas.obter_visita_aberta(sender)
            assert final.startswith("Visita aberta.")
            assert visita["estado_fluxo"] == "visita_aberta"
            assert visita["fazenda"] == "Fazenda Imperial"
            assert visita["proprietario"] == "Alexander Duarte Paniago"
            assert visita["gerente"] == "Paulo Silva"
            assert visita["area_hectares"] == 2299
            assert visita["safra"] == "26/27"
            assert visita["tipo_visita"] == "Apresentacao de produtos ao cliente"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_comandos_diretos():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            api_whatsapp.handle_rdv_text_message(sender, "visita")
            visita = visitas.obter_visita_aberta(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            api_whatsapp.handle_rdv_text_message(sender, "fazenda Fazenda Imperial")
            api_whatsapp.handle_rdv_text_message(sender, "gerente Paulo Silva")
            api_whatsapp.handle_rdv_text_message(sender, "hectares 2299")
            api_whatsapp.handle_rdv_text_message(sender, "obs Pedido de 300T")

            saved = visitas.obter_visita_aberta(sender)
            assert saved["fazenda"] == "Fazenda Imperial"
            assert saved["gerente"] == "Paulo Silva"
            assert saved["area_hectares"] == 2299
            assert "Pedido de 300T" in saved["observacoes"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_salvar_localizacao():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            reply = api_whatsapp.handle_visitas_location_message(
                sender,
                {"latitude": -15.0019124, "longitude": -50.7714295},
            )

            saved = visitas.obter_visita_aberta(sender)
            assert "📍 Localização salva." in reply
            assert "https://maps.google.com/?q=-15.0019124,-50.7714295" in reply
            assert saved["maps_url_principal"] == (
                "https://maps.google.com/?q=-15.0019124,-50.7714295"
            )
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_fechar():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "fechar visita")

            closed = visitas.obter_visita(visita["id"])
            assert "Prévia do relatório enviada" in reply
            assert "A visita ainda não foi finalizada" in reply
            assert sent[0][4].startswith(b"%PDF")
            assert closed["status"] == "aberta"
            assert closed["estado_fluxo"] == "aguardando_revisao_final"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_visita_finalizada_nao_recebe_nova_midia():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            api_whatsapp.handle_rdv_text_message(sender, "1")
            photo_reply = api_whatsapp.handle_visitas_media_message(
                sender,
                "image",
                "media-after-close",
                "foto-after-close.jpg",
            )
            video_reply = api_whatsapp.handle_visitas_video_message(
                sender,
                "video-after-close",
                "video/mp4",
            )
            closed = visitas.obter_visita_completa(visita["id"])

            assert "Esta visita já foi finalizada" in photo_reply
            assert "Esta visita já foi finalizada" in video_reply
            assert closed["status"] == "fechada"
            assert closed["midias"] == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_visita_revisao_corrige_descricao_e_gera_nova_previa_sem_fechar():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            visitas.atualizar_campo(visita["id"], "descricao_visita", "Descricao antiga da visita")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            edit_prompt = api_whatsapp.handle_rdv_text_message(sender, "3")
            updated = api_whatsapp.handle_rdv_text_message(
                sender,
                "Descricao atualizada com detalhes suficientes",
            )
            preview = api_whatsapp.handle_rdv_text_message(sender, "previa")
            saved = visitas.obter_visita(visita["id"])

            assert "nova descricao" in edit_prompt.lower()
            assert "Informação atualizada" in updated
            assert "prévia anterior pode estar desatualizada" in updated.lower()
            assert "Prévia do relatório enviada" in preview
            assert len(sent) == 2
            assert saved["descricao_visita"] == "Descricao atualizada com detalhes suficientes"
            assert saved["status"] == "aberta"
            assert saved["estado_fluxo"] == "aguardando_revisao_final"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_visita_revisao_corrige_observacoes_e_finaliza_por_texto():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            api_whatsapp.handle_rdv_text_message(sender, "encerrar visita")
            api_whatsapp.handle_rdv_text_message(sender, "4")
            api_whatsapp.handle_rdv_text_message(sender, "1")
            updated = api_whatsapp.handle_rdv_text_message(sender, "Observacao revisada")
            final = api_whatsapp.handle_rdv_text_message(sender, "finalizar")
            closed = visitas.obter_visita(visita["id"])

            assert "Informação atualizada" in updated
            assert "Visita finalizada com sucesso" in final
            assert closed["observacoes_gerais"] == "Observacao revisada"
            assert closed["status"] == "fechada"
            assert closed["fechado_em"]
            assert len(sent) == 2
            assert all(item[1] == f"relatorio_visita_{visita['id']}.pdf" for item in sent)
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_visita_revisao_corrige_dados_da_propriedade():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            menu = api_whatsapp.handle_rdv_text_message(sender, "2")
            prompt = api_whatsapp.handle_rdv_text_message(sender, "7")
            updated = api_whatsapp.handle_rdv_text_message(sender, "500 hectares")
            saved = visitas.obter_visita(visita["id"])

            assert "Localização da fazenda" in menu
            assert "tamanho total" in prompt.lower()
            assert "Informação atualizada" in updated
            assert saved["area"] == "500 hectares"
            assert saved["estado_fluxo"] == "aguardando_revisao_final"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_visita_revisao_aceita_localizacao_nativa_e_link_maps():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            native_reply = api_whatsapp.handle_visitas_location_message(
                sender,
                {"latitude": -15.0019124, "longitude": -50.7714295},
            )
            link_reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "https://maps.google.com/?q=-16,-51",
            )
            saved = visitas.obter_visita_completa(visita["id"])

            assert "Localização atualizada" in native_reply
            assert "prévia anterior pode estar desatualizada" in native_reply.lower()
            assert "Localização atualizada" in link_reply
            assert saved["maps_url_principal"] == "https://maps.google.com/?q=-16,-51"
            assert saved["localizacao_texto"] == "https://maps.google.com/?q=-16,-51"
            assert saved["status"] == "aberta"
            assert saved["estado_fluxo"] == "aguardando_revisao_final"
            assert len(saved["localizacoes"]) == 1
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_visita_revisao_previa_apos_midia_reenvia_pdf_sem_fechar():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            api_whatsapp.handle_visitas_media_message(
                sender,
                "image",
                "media-review",
                "foto-review.jpg",
            )
            api_whatsapp.handle_rdv_text_message(sender, "2")
            reply = api_whatsapp.handle_rdv_text_message(sender, "previa")
            saved = visitas.obter_visita_completa(visita["id"])

            assert "Prévia do relatório enviada" in reply
            assert len(sent) == 2
            assert sent[-1][4].startswith(b"%PDF")
            assert len(saved["midias"]) == 1
            assert saved["status"] == "aberta"
            assert saved["estado_fluxo"] == "aguardando_revisao_final"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_nova_visita_nao_cria_segunda_enquanto_existe_aberta():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            antiga = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(antiga["id"], "fazenda", "FAZENDA NOVA FRONTEIRA")
            visitas.atualizar_campo(antiga["id"], "estado_fluxo", "visita_aberta")

            start = api_whatsapp.handle_rdv_text_message(sender, "nova visita")
            created = api_whatsapp.handle_rdv_text_message(sender, "Itapuã Prestes")
            data = visitas.listar_visitas_validas()
            abertas = [item for item in data["visitas"] if item["status"] == "aberta"]
            assert "Visita em andamento" in start
            assert "Visita em andamento" in created
            assert len(abertas) == 1
            assert abertas[0]["id"] == antiga["id"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_active_states.clear()
        api_whatsapp.visita_new_visit_states.clear()


def test_visita_com_aberta_existente_pede_escolha():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "FAZENDA NOVA FRONTEIRA")

            reply = api_whatsapp.handle_rdv_text_message(sender, "visita")

            assert "Visita em andamento" in reply
            assert f"continuar visita {visita['id']}" in reply
            assert "Revisar e finalizar" in reply
            assert "fechar visita" in reply
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_active_states.clear()
        api_whatsapp.visita_new_visit_states.clear()


def test_continuar_visita_define_visita_ativa():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            start = api_whatsapp.handle_rdv_text_message(sender, f"continuar visita {visita['id']}")
            reply = api_whatsapp.handle_visitas_media_message(
                sender,
                "image",
                "media-2",
                "foto-imperial.jpg",
            )

            saved = visitas.obter_visita_completa(visita["id"])
            assert "Você voltou para a visita" in start
            assert "Foto salva na visita Fazenda Imperial." in reply
            assert len(saved["midias"]) == 1
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_active_states.clear()


def test_visita_revisao_final_aceita_foto_e_comentario_sem_finalizar():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            api_whatsapp.handle_rdv_text_message(sender, "nova visita")
            api_whatsapp.handle_rdv_text_message(sender, "Itapuã Prestes")
            visita_id = api_whatsapp.visita_active_states[sender]
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None

            close_reply = api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            media_reply = api_whatsapp.handle_visitas_media_message(
                sender,
                "image",
                "media-3",
                "foto-fechada.jpg",
            )

            saved = visitas.obter_visita_completa(visita_id)
            assert "Prévia do relatório enviada" in close_reply
            assert "Foto anexada" in media_reply
            assert "prévia anterior pode estar desatualizada" in media_reply.lower()
            assert "Digite o comentario da Foto" in api_whatsapp.handle_rdv_text_message(sender, "1")
            done = api_whatsapp.handle_rdv_text_message(sender, "Foto complementar")
            saved = visitas.obter_visita_completa(visita_id)
            assert "Comentário salvo" in done
            assert "prévia anterior pode estar desatualizada" in done.lower()
            assert saved["status"] == "aberta"
            assert saved["estado_fluxo"] == "aguardando_revisao_final"
            assert len(saved["midias"]) == 1
            assert saved["midias"][0]["comentario"] == "Foto complementar"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_active_states.clear()
        api_whatsapp.visita_new_visit_states.clear()


def test_visita_revisao_final_apaga_foto_com_confirmacao_e_pdf_atualizado():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            foto1_path = Path(temp_dir) / "foto-1.jpg"
            foto2_path = Path(temp_dir) / "foto-2.jpg"
            foto1_path.write_bytes(b"foto-1")
            foto2_path.write_bytes(b"foto-2")
            foto1 = visitas.adicionar_midia(visita["id"], "foto", caminho_arquivo=str(foto1_path))
            foto2 = visitas.adicionar_midia(visita["id"], "foto", caminho_arquivo=str(foto2_path))
            visitas.salvar_comentario_foto(foto1["id"], "Comentario foto mantida")
            visitas.salvar_comentario_foto(foto2["id"], "Comentario foto removida")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            menu = api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            first_list = api_whatsapp.handle_rdv_text_message(sender, "6")
            invalid = api_whatsapp.handle_rdv_text_message(sender, "9")
            canceled = api_whatsapp.handle_rdv_text_message(sender, "cancelar")
            api_whatsapp.handle_rdv_text_message(sender, "apagar foto")
            confirm_prompt = api_whatsapp.handle_rdv_text_message(sender, "2")
            denied = api_whatsapp.handle_rdv_text_message(sender, "2")
            api_whatsapp.handle_rdv_text_message(sender, "apagar foto")
            api_whatsapp.handle_rdv_text_message(sender, "2")
            removed = api_whatsapp.handle_rdv_text_message(sender, "1")
            preview = api_whatsapp.handle_rdv_text_message(sender, "5")
            saved = visitas.obter_visita_completa(visita["id"])
            pdf_text = _extract_pdf_text(sent[-1][4])

            assert "6. Apagar foto" in menu
            assert "7. Apagar v" in menu
            assert "Fotos da visita" in first_list
            assert "Comentario foto removida" in first_list
            assert "Número inválido" in invalid
            assert "A visita ainda não foi finalizada" in canceled
            assert "Deseja apagar a Foto 2" in confirm_prompt
            assert "A visita ainda não foi finalizada" in denied
            assert "Foto 2 removida da visita" in removed
            assert "prévia anterior pode estar desatualizada" in removed
            assert "relat" in preview
            assert len(saved["midias"]) == 1
            assert saved["midias"][0]["comentario"] == "Comentario foto mantida"
            assert foto1_path.exists()
            assert not foto2_path.exists()
            assert "Quantidade de fotos" in pdf_text
            assert "1" in pdf_text
            assert "Comentario foto mantida" in pdf_text
            assert "Comentario foto removida" not in pdf_text
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_active_states.clear()


def test_visita_revisao_final_envia_menu_interativo_e_parser_list_reply(monkeypatch):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None
            sent_lists = []
            monkeypatch.setattr(
                api_whatsapp,
                "send_whatsapp_list_message",
                lambda **kwargs: sent_lists.append(kwargs),
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            api_whatsapp._send_rdv_reply(sender, "fechar visita", reply)
            rows = sent_lists[0]["sections"][0]["rows"]
            command = api_whatsapp._extract_text(
                {
                    "type": "interactive",
                    "interactive": {
                        "list_reply": {
                            "id": "visita_revisao_apagar_foto",
                            "title": "Apagar foto",
                        }
                    },
                }
            )

            assert sent_lists[0]["header"] == "Revisar visita"
            assert [row["title"] for row in rows] == [
                "Finalizar visita",
                "Corrigir dados",
                "Corrigir descrição",
                "Corrigir observações",
                "Gerar nova prévia",
                "Apagar foto",
                "Apagar vídeo",
                "Voltar sem finalizar",
            ]
            assert command == "6"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_active_states.clear()


def test_visita_revisao_final_menu_interativo_fallback_textual(monkeypatch):
    sent_texts = []
    error = api_whatsapp.WhatsAppSendError(
        category="INVALID_PAYLOAD",
        fallback_allowed=True,
        message_kind="interactive.list",
    )
    monkeypatch.setattr(
        api_whatsapp,
        "send_whatsapp_list_message",
        lambda **kwargs: (_ for _ in ()).throw(error),
    )
    monkeypatch.setattr(
        api_whatsapp,
        "send_whatsapp_text",
        lambda to, message: sent_texts.append((to, message)),
    )

    reply = api_whatsapp._visita_revisao_final_message()
    api_whatsapp._send_rdv_reply("5500000000001", "fechar visita", reply)

    assert sent_texts == [("5500000000001", reply)]


def test_visita_revisao_final_menu_interativo_timeout_nao_faz_fallback_textual(monkeypatch):
    sent_texts = []
    error = api_whatsapp.WhatsAppSendError(
        category="TIMEOUT",
        retryable=True,
        fallback_allowed=False,
        message_kind="interactive.list",
    )
    monkeypatch.setattr(
        api_whatsapp,
        "send_whatsapp_list_message",
        lambda **kwargs: (_ for _ in ()).throw(error),
    )
    monkeypatch.setattr(
        api_whatsapp,
        "send_whatsapp_text",
        lambda to, message: sent_texts.append((to, message)),
    )

    reply = api_whatsapp._visita_revisao_final_message()
    with pytest.raises(api_whatsapp.WhatsAppSendError):
        api_whatsapp._send_rdv_reply("5500000000001", "fechar visita", reply)

    assert sent_texts == []


def test_visita_apagar_foto_usa_botoes_e_confirma_sim_por_interactive(monkeypatch):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            foto1 = visitas.adicionar_midia(visita["id"], "foto", caminho_arquivo=str(Path(temp_dir) / "foto-1.jpg"))
            foto2 = visitas.adicionar_midia(visita["id"], "foto", caminho_arquivo=str(Path(temp_dir) / "foto-2.jpg"))
            visitas.salvar_comentario_foto(foto1["id"], "Foto mantida")
            visitas.salvar_comentario_foto(foto2["id"], "Foto removida")
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None
            sent_buttons = []
            monkeypatch.setattr(
                api_whatsapp,
                "send_whatsapp_button_message",
                lambda **kwargs: sent_buttons.append(kwargs),
            )

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            list_reply = api_whatsapp.handle_rdv_text_message(sender, "6")
            api_whatsapp._send_rdv_reply(sender, "6", list_reply)
            choose_command = api_whatsapp._extract_text(
                {
                    "type": "interactive",
                    "interactive": {
                        "button_reply": {
                            "id": "visita_apagar_foto_2",
                            "title": "Foto 2",
                        }
                    },
                }
            )
            confirm_reply = api_whatsapp.handle_rdv_text_message(sender, choose_command)
            api_whatsapp._send_rdv_reply(sender, choose_command, confirm_reply)
            yes_command = api_whatsapp._extract_text(
                {
                    "type": "interactive",
                    "interactive": {
                        "button_reply": {
                            "id": "visita_confirmar_apagar_midia_sim",
                            "title": "Sim",
                        }
                    },
                }
            )
            removed = api_whatsapp.handle_rdv_text_message(sender, yes_command)
            saved = visitas.obter_visita_completa(visita["id"])

            assert [button["title"] for button in sent_buttons[0]["buttons"]] == ["Foto 1", "Foto 2", "Cancelar"]
            assert [button["title"] for button in sent_buttons[1]["buttons"]] == ["Sim", "Não"]
            assert "Foto 2 removida da visita" in removed
            assert len(saved["midias"]) == 1
            assert saved["midias"][0]["comentario"] == "Foto mantida"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_active_states.clear()


def test_visita_apagar_video_usa_lista_e_confirma_nao_por_interactive(monkeypatch):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            for index in range(1, 4):
                video = visitas.adicionar_midia(
                    visita["id"],
                    "video",
                    storage_key=f"visitas/2026/07/08/1/videos/video-{index}.mp4",
                    public_url=f"https://cdn.example/video-{index}.mp4",
                    tamanho_bytes=1024 * index,
                    mime_type="video/mp4",
                )
                visitas.salvar_comentario_midia(video["id"], f"Legenda {index}")
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None
            sent_lists = []
            sent_buttons = []
            monkeypatch.setattr(
                api_whatsapp,
                "send_whatsapp_list_message",
                lambda **kwargs: sent_lists.append(kwargs),
            )
            monkeypatch.setattr(
                api_whatsapp,
                "send_whatsapp_button_message",
                lambda **kwargs: sent_buttons.append(kwargs),
            )

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            list_reply = api_whatsapp.handle_rdv_text_message(sender, "7")
            api_whatsapp._send_rdv_reply(sender, "7", list_reply)
            choose_command = api_whatsapp._extract_text(
                {
                    "type": "interactive",
                    "interactive": {
                        "list_reply": {
                            "id": "visita_apagar_video_2",
                            "title": "Vídeo 2",
                        }
                    },
                }
            )
            confirm_reply = api_whatsapp.handle_rdv_text_message(sender, choose_command)
            api_whatsapp._send_rdv_reply(sender, choose_command, confirm_reply)
            no_command = api_whatsapp._extract_text(
                {
                    "type": "interactive",
                    "interactive": {
                        "button_reply": {
                            "id": "visita_confirmar_apagar_midia_nao",
                            "title": "Não",
                        }
                    },
                }
            )
            denied = api_whatsapp.handle_rdv_text_message(sender, no_command)
            saved = visitas.obter_visita_completa(visita["id"])

            rows = sent_lists[0]["sections"][0]["rows"]
            assert [row["title"] for row in rows] == ["Vídeo 1", "Vídeo 2", "Vídeo 3", "Cancelar"]
            assert [button["title"] for button in sent_buttons[0]["buttons"]] == ["Sim", "Não"]
            assert "A visita ainda" in denied
            assert len(saved["midias"]) == 3
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_active_states.clear()


def test_visita_revisao_final_apaga_video_com_confirmacao_storage_e_pdf_atualizado(monkeypatch):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            video1 = visitas.adicionar_midia(
                visita["id"],
                "video",
                storage_key="visitas/2026/07/08/1/videos/video-1.mp4",
                public_url="https://cdn.example/video-1.mp4",
                tamanho_bytes=2048,
                mime_type="video/mp4",
            )
            video2 = visitas.adicionar_midia(
                visita["id"],
                "video",
                storage_key="visitas/2026/07/08/1/videos/video-2.mp4",
                public_url="https://cdn.example/video-2.mp4",
                tamanho_bytes=4096,
                mime_type="video/mp4",
            )
            visitas.salvar_comentario_midia(video1["id"], "Legenda mantida")
            visitas.salvar_comentario_midia(video2["id"], "Legenda removida")
            sent = []
            deleted_keys = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )
            monkeypatch.setattr(
                api_whatsapp,
                "delete_storage_file",
                lambda storage_key: deleted_keys.append(storage_key),
            )

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            listed = api_whatsapp.handle_rdv_text_message(sender, "7")
            confirm_prompt = api_whatsapp.handle_rdv_text_message(sender, "2")
            removed = api_whatsapp.handle_rdv_text_message(sender, "sim")
            api_whatsapp.handle_rdv_text_message(sender, "previa")
            saved = visitas.obter_visita_completa(visita["id"])
            pdf_text = _extract_pdf_text(sent[-1][4])

            assert "V" in listed
            assert "Legenda removida" in listed
            assert "Deseja apagar a V" in confirm_prompt
            assert "2" in confirm_prompt
            assert "removido da visita" in removed
            assert deleted_keys == ["visitas/2026/07/08/1/videos/video-2.mp4"]
            assert len(saved["midias"]) == 1
            assert saved["midias"][0]["comentario"] == "Legenda mantida"
            assert "Quantidade de v" in pdf_text
            assert "Legenda mantida" in pdf_text
            assert "Legenda removida" not in pdf_text
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_active_states.clear()


def test_visita_revisao_final_mensagens_sem_midia_e_bloqueia_finalizada():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None

            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            no_photo = api_whatsapp.handle_rdv_text_message(sender, "6")
            no_video = api_whatsapp.handle_rdv_text_message(sender, "7")
            api_whatsapp.handle_rdv_text_message(sender, "1")
            blocked = api_whatsapp.handle_rdv_text_message(sender, "apagar foto")

            assert "não possui fotos" in no_photo
            assert "não possui v" in no_video
            assert "já foi finalizada" in blocked
            assert "Não é possível apagar mídias" in blocked
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_active_states.clear()


def test_cancelar_visita_limpa_estado_ativo():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            api_whatsapp.handle_rdv_text_message(sender, "nova visita")
            api_whatsapp.handle_rdv_text_message(sender, "Itapuã Prestes")
            visita_id = api_whatsapp.visita_active_states[sender]

            cancel_reply = api_whatsapp.handle_rdv_text_message(sender, "cancelar visita")
            media_reply = api_whatsapp.handle_visitas_media_message(
                sender,
                "image",
                "media-4",
                "foto-cancelada.jpg",
            )

            saved = visitas.obter_visita_completa(visita_id)
            assert cancel_reply == "Visita cancelada com sucesso."
            assert media_reply == "Nenhuma visita em andamento encontrada."
            assert saved["midias"] == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_active_states.clear()
        api_whatsapp.visita_new_visit_states.clear()


def test_fluxo_edicao_nao_interfere_com_visita_ativa():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            api_whatsapp.handle_rdv_text_message(sender, "nova visita")
            api_whatsapp.handle_rdv_text_message(sender, "Itapuã Prestes")
            visita_id = api_whatsapp.visita_active_states[sender]
            visitas.atualizar_campo(visita_id, "gerente", "Marcos")

            api_whatsapp.handle_rdv_text_message(sender, f"editar visita {visita_id}")
            reply = api_whatsapp.handle_rdv_text_message(sender, "gerente = X")

            saved = visitas.obter_visita_completa(visita_id)
            assert "Campo atualizado:" in reply
            assert saved["gerente"] == "X"
            assert saved["dados_coletados"] == []
            assert saved["observacoes"] in (None, "")
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_edit_states.clear()
        api_whatsapp.visita_active_states.clear()


def test_ver_visita_por_id():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Marcelo Fukuda")
            visitas.atualizar_campo(visita["id"], "fazenda", "FAZENDA NOVA FRONTEIRA")
            visitas.atualizar_campo(visita["id"], "gerente", "Marcos")
            visitas.atualizar_campo(visita["id"], "data_visita", "2026-06-17")

            reply = api_whatsapp.handle_rdv_text_message(sender, f"ver visita {visita['id']}")

            assert f"Visita #{visita['id']} - FAZENDA NOVA FRONTEIRA" in reply
            assert "Status: aberta" in reply
            assert "Técnico: Marcelo Fukuda" in reply
            assert "Data: 17/06/2026" in reply
            assert "Gerente/responsável: Marcos" in reply
            assert f"editar visita {visita['id']}" in reply
            assert f"relatório visita {visita['id']}" in reply
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_edit_states.clear()


def test_editar_visita_campo_gerente():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            visitas.atualizar_campo(visita["id"], "gerente", "Marcos")

            start = api_whatsapp.handle_rdv_text_message(sender, f"editar visita {visita['id']}")
            reply = api_whatsapp.handle_rdv_text_message(sender, "gerente = Marcos Silva")

            saved = visitas.obter_visita(visita["id"])
            assert f"Você está editando a visita #{visita['id']}" in start
            assert saved["gerente"] == "Marcos Silva"
            assert "Campo atualizado:" in reply
            assert "Gerente/responsável" in reply
            assert "Antes: Marcos" in reply
            assert "Depois: Marcos Silva" in reply
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_edit_states.clear()


def test_editar_visita_observacoes_reflete_no_pdf():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            api_whatsapp.handle_rdv_text_message(sender, f"editar visita {visita['id']}")
            api_whatsapp.handle_rdv_text_message(
                sender,
                "observações = Cliente solicitou orçamento.",
            )
            reply = api_whatsapp.handle_rdv_text_message(sender, f"relatorio visita {visita['id']}")

            saved = visitas.obter_visita_completa(visita["id"])
            assert reply is None
            assert saved["observacoes"] == "Cliente solicitou orçamento."
            assert sent[0][4].startswith(b"%PDF")
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_edit_states.clear()


def test_editar_visita_cancelada_bloqueia():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.cancelar_visita(visita["id"])

            reply = api_whatsapp.handle_rdv_text_message(sender, f"editar visita {visita['id']}")

            assert "cancelada" in reply
            assert "não pode ser editada" in reply
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_edit_states.clear()


def test_editar_visita_campo_invalido():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")

            api_whatsapp.handle_rdv_text_message(sender, f"editar visita {visita['id']}")
            reply = api_whatsapp.handle_rdv_text_message(sender, "campo_invalido = x")

            assert "Não reconheci esse campo." in reply
            assert "Campos aceitos" in reply
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_edit_states.clear()


def test_fechar_edicao():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "gerente", "Marcos")
            visitas.fechar_visita(visita["id"])

            api_whatsapp.handle_rdv_text_message(sender, f"editar visita {visita['id']}")
            close_reply = api_whatsapp.handle_rdv_text_message(sender, "fechar edição")
            next_reply = api_whatsapp.handle_rdv_text_message(sender, "gerente = Marcos Silva")

            saved = visitas.obter_visita(visita["id"])
            assert "Edição finalizada." in close_reply
            assert saved["gerente"] == "Marcos"
            assert next_reply == api_whatsapp.MAIN_MENU_MESSAGE
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_edit_states.clear()


def test_cancelar_edicao():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")

            api_whatsapp.handle_rdv_text_message(sender, f"editar visita {visita['id']}")
            reply = api_whatsapp.handle_rdv_text_message(sender, "cancelar edição")

            assert "Edição encerrada." in reply
            assert api_whatsapp.visita_edit_states == {}
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_edit_states.clear()


def test_edicao_de_outro_telefone_permitida():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            rdv, visitas, sender = _install_services(temp_dir)
            other = rdv.get_collaborator_by_phone("5500000000002")["telefone_whatsapp"]
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "gerente", "Marcos")

            start = api_whatsapp.handle_rdv_text_message(other, f"editar visita {visita['id']}")
            reply = api_whatsapp.handle_rdv_text_message(other, "gerente = Marcos Silva")

            assert "Você está editando" in start
            assert "Depois: Marcos Silva" in reply
            assert visitas.obter_visita(visita["id"])["gerente"] == "Marcos Silva"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_edit_states.clear()


def test_relatorio_visita_envia_pdf():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            visitas.atualizar_campo(visita["id"], "gerente", "Paulo Silva")
            visitas.adicionar_observacao(visita["id"], "Pedido de 300T")
            sent = []

            def fake_send(to, content, filename, caption, mime_type):
                sent.append(
                    {
                        "to": to,
                        "content": content,
                        "filename": filename,
                        "caption": caption,
                        "mime_type": mime_type,
                    }
                )

            api_whatsapp.send_whatsapp_document = fake_send

            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                f"relatorio visita {visita['id']}",
            )

            assert reply is None
            assert len(sent) == 1
            assert sent[0]["to"] == sender
            assert sent[0]["content"].startswith(b"%PDF")
            assert sent[0]["filename"] == f"relatorio_visita_{visita['id']}.pdf"
            assert sent[0]["caption"] == (
                "Segue o relatório da visita técnica da Ciclus Agro."
            )
            assert sent[0]["mime_type"] == "application/pdf"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_relatorio_visita_sem_visita():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, _, sender = _install_services(temp_dir)

            reply = api_whatsapp.handle_rdv_text_message(sender, "relatorio visita")

            assert reply == "Você ainda não possui relatórios de visitas finalizadas."
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_relatorio_visita_por_id_de_outro_telefone_nao_expoe_pdf():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            rdv, visitas, sender = _install_services(temp_dir)
            other = rdv.get_collaborator_by_phone("5500000000002")["telefone_whatsapp"]
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            visitas.fechar_visita(visita["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(
                other,
                f"relatorio visita {visita['id']}",
            )

            assert "Não encontrei" in reply
            assert sent == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_pdf_visita_por_id_envia_pdf_individual():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            assert api_whatsapp.handle_rdv_text_message(
                sender,
                f"pdf visita {visita['id']}",
            ) is None
            assert api_whatsapp.handle_rdv_text_message(
                sender,
                f"visita pdf {visita['id']}",
            ) is None

            assert len(sent) == 2
            assert sent[0][1] == f"relatorio_visita_{visita['id']}.pdf"
            assert sent[1][1] == f"relatorio_visita_{visita['id']}.pdf"
            assert sent[0][4].startswith(b"%PDF")
            assert sent[1][4].startswith(b"%PDF")
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_pdf_visita_sem_id_lista_visitas_sem_gerar_pdf():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            visitas.fechar_visita(visita["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "pdf visita")

            assert "Escolha uma pelo ID" in reply
            assert f"#{visita['id']} - Fazenda Imperial" in reply
            assert f"relatório visita {visita['id']}" in reply
            assert sent == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_relatorio_visita_sem_id_com_multiplas_visitas_lista_opcoes():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            rdv, visitas, sender = _install_services(temp_dir)
            primeira = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(primeira["id"], "fazenda", "Fazenda Imperial")
            visitas.fechar_visita(primeira["id"])
            segunda = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(segunda["id"], "fazenda", "Fazenda Boi Dourado 3J")
            visitas.fechar_visita(segunda["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "relatorio visita")

            assert "Escolha uma pelo ID" in reply
            assert f"#{primeira['id']} - Fazenda Imperial" in reply
            assert f"#{segunda['id']} - Fazenda Boi Dourado 3J" in reply
            assert sent == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_relatorio_fazenda_um_resultado():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            visitas.fechar_visita(visita["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "relatorio fazenda Fazenda Imperial",
            )

            assert reply is None
            assert len(sent) == 1
            assert sent[0][1] == f"relatorio_visita_{visita['id']}.pdf"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_relatorio_fazenda_multiplos_resultados():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            rdv, visitas, sender = _install_services(temp_dir)
            primeira = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(primeira["id"], "fazenda", "Fazenda Imperial")
            visitas.fechar_visita(primeira["id"])
            segunda = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(segunda["id"], "fazenda", "Fazenda Imperial Norte")
            visitas.fechar_visita(segunda["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "relatorio fazenda Fazenda Imperial",
            )

            assert "Encontrei mais de uma visita" in reply
            assert f"#{primeira['id']} - Fazenda Imperial" in reply
            assert f"#{segunda['id']} - Fazenda Imperial Norte" in reply
            assert sent == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_relatorio_visita_nao_gera_pdf_de_cancelada():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Cancelada")
            visitas.cancelar_visita(visita["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "relatório visita")

            assert reply == "Você ainda não possui relatórios de visitas finalizadas."
            assert sent == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_relatorio_visita_id_cancelado_bloqueia():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Cancelada")
            visitas.cancelar_visita(visita["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                f"relatório visita {visita['id']}",
            )

            assert reply == api_whatsapp.CANCELED_VISITA_REPORT_MESSAGE
            assert sent == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_fazendas_visitadas_exclui_canceladas():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            cancelada = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(cancelada["id"], "data_visita", "2026-06-17")
            visitas.atualizar_campo(cancelada["id"], "fazenda", "Fazenda Cancelada")
            visitas.cancelar_visita(cancelada["id"])
            fechada = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(fechada["id"], "data_visita", "2026-06-17")
            visitas.atualizar_campo(fechada["id"], "fazenda", "Fazenda Valida")
            visitas.fechar_visita(fechada["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "fazendas visitadas")

            assert reply is None
            workbook = load_workbook(BytesIO(sent[0][4]))
            visitas_sheet = workbook["Visitas"]
            fazendas = [
                row[0]
                for row in visitas_sheet.iter_rows(
                    min_row=2,
                    min_col=5,
                    max_col=5,
                    values_only=True,
                )
            ]
            assert "Fazenda Valida" in fazendas
            assert "Fazenda Cancelada" not in fazendas
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_listar_visitas_textual_inclui_mais_de_dez_sem_deduplicar_fazenda():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, _sender = _install_services(temp_dir)
            created = []
            for index in range(1, 19):
                created.append(
                    _create_test_visit(
                        visitas,
                        index,
                        fazenda="Fazenda Repetida" if index in {3, 14} else f"Fazenda {index:02d}",
                        data_visita=f"2026-07-{index:02d}",
                        status="aberta" if index % 5 == 0 else "fechada",
                    )
                )

            messages = api_whatsapp._listar_visitas_messages("visitas")
            text = "\n\n".join(messages)
            ids = _ids_in_message_text(text)

            assert "Visitas tÃ©cnicas encontradas: 18" in messages[0]
            assert len(ids) == 18
            assert len(set(ids)) == 18
            assert ids[0] == created[-1]["id"]
            assert ids[-1] == created[0]["id"]
            assert text.count("Fazenda Repetida") == 2
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_listar_visitas_textual_nao_envia_limite_fixo(monkeypatch):
    calls = []

    class FakeVisitasService:
        def listar_visitas_validas(self, **kwargs):
            calls.append(kwargs)
            return {"visitas": []}

    monkeypatch.setattr(api_whatsapp, "visitas_service", FakeVisitasService())

    assert api_whatsapp._listar_visitas_messages("visitas") == [
        api_whatsapp.NO_VALID_VISITA_MESSAGE
    ]
    assert calls == [{}]


def test_listar_visitas_abertas_e_hoje_sem_limite():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, _sender = _install_services(temp_dir)
            today = api_whatsapp.date.today().isoformat()
            abertas = [
                _create_test_visit(visitas, index, data_visita=today, status="aberta")
                for index in range(1, 13)
            ]
            fechada_hoje = _create_test_visit(
                visitas,
                30,
                data_visita=today,
                status="fechada",
                fazenda="Fazenda Fechada Hoje",
            )
            _create_test_visit(
                visitas,
                31,
                data_visita="2026-01-01",
                status="fechada",
                fazenda="Fazenda Antiga",
            )

            abertas_text = "\n\n".join(api_whatsapp._listar_visitas_messages("visitas abertas"))
            hoje_text = "\n\n".join(api_whatsapp._listar_visitas_messages("visitas hoje"))
            abertas_ids = _ids_in_message_text(abertas_text)
            hoje_ids = _ids_in_message_text(hoje_text)

            assert "Visitas abertas encontradas: 12" in abertas_text
            assert len(abertas_ids) == 12
            assert fechada_hoje["id"] not in abertas_ids
            assert "Visitas tÃ©cnicas encontradas hoje: 13" in hoje_text
            assert len(hoje_ids) == 13
            assert fechada_hoje["id"] in hoje_ids
            assert all(item["id"] in hoje_ids for item in abertas)
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_listar_visitas_sem_resultado_preserva_mensagem():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _install_services(temp_dir)
            assert api_whatsapp._listar_visitas_messages("visitas") == [
                api_whatsapp.NO_VALID_VISITA_MESSAGE
            ]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_listar_visitas_pagina_por_tamanho_sem_cortar_blocos_e_instrucoes_so_no_final():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, _sender = _install_services(temp_dir)
            for index in range(1, 31):
                _create_test_visit(
                    visitas,
                    index,
                    fazenda=f"Fazenda Com Nome Muito Longo Para Testar Paginacao {index:02d}",
                    data_visita=f"2026-07-{index:02d}",
                    gerente=f"Gerente Com Nome Longo {index:02d}",
                )

            messages = api_whatsapp._listar_visitas_messages("visitas")
            text = "\n\n".join(messages)
            ids = _ids_in_message_text(text)

            assert len(messages) > 1
            assert all(len(message) <= api_whatsapp.VISITA_LIST_MESSAGE_MAX_CHARS for message in messages)
            assert len(ids) == 30
            assert len(set(ids)) == 30
            assert ids == sorted(ids, reverse=True)
            for message in messages:
                for block in message.split("\n\n"):
                    if block.startswith("#"):
                        assert "Status:" in block
                        assert "Gerente:" in block
            assert sum("Para gerar PDF individual" in message for message in messages) == 1
            assert sum("Para buscar por fazenda" in message for message in messages) == 1
            assert "Para gerar PDF individual" in messages[-1]
            assert "Para buscar por fazenda" in messages[-1]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_listar_visitas_fluxo_envia_cada_parte_por_texto(monkeypatch):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            for index in range(1, 31):
                _create_test_visit(
                    visitas,
                    index,
                    fazenda=f"Fazenda Longa Para Envio Parcelado {index:02d}",
                    data_visita=f"2026-07-{index:02d}",
                )
            sent = []
            monkeypatch.setattr(
                api_whatsapp,
                "send_whatsapp_text",
                lambda to, message: sent.append((to, message)),
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "visitas")

            expected = api_whatsapp._listar_visitas_messages("visitas")
            assert reply is None
            assert sent == [(sender, message) for message in expected]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_listar_visitas_fluxo_propaga_erro_meta_sem_retry_ou_fallback(monkeypatch):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            for index in range(1, 31):
                _create_test_visit(
                    visitas,
                    index,
                    fazenda=f"Fazenda Longa Para Falha Parcial {index:02d}",
                    data_visita=f"2026-07-{index:02d}",
                )
            calls = []
            error = api_whatsapp.WhatsAppSendError(
                category="TIMEOUT",
                retryable=True,
                fallback_allowed=False,
                message_kind="text",
            )

            def fake_send(to, message):
                calls.append((to, message))
                if len(calls) == 2:
                    raise error

            monkeypatch.setattr(api_whatsapp, "send_whatsapp_text", fake_send)

            with pytest.raises(api_whatsapp.WhatsAppSendError):
                api_whatsapp.handle_rdv_text_message(sender, "visitas")

            assert len(calls) == 2
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_planilha_visitas_global():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            primeira = visitas.iniciar_visita("5500000000001", tecnico_nome="Danilo")
            visitas.atualizar_campo(primeira["id"], "data_visita", "2026-05-10")
            visitas.atualizar_campo(primeira["id"], "fazenda", "Fazenda Imperial")
            segunda = visitas.iniciar_visita("5500000000002", tecnico_nome="Marcelo")
            visitas.atualizar_campo(segunda["id"], "data_visita", "2026-06-17")
            visitas.atualizar_campo(segunda["id"], "fazenda", "Fazenda Boi Dourado 3J")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "planilha visitas")

            assert reply is None
            workbook = load_workbook(BytesIO(sent[0][4]))
            sheet = workbook["Visitas"]
            fazendas = [
                row[0]
                for row in sheet.iter_rows(
                    min_row=2,
                    min_col=5,
                    max_col=5,
                    values_only=True,
                )
            ]
            assert "Fazenda Imperial" in fazendas
            assert "Fazenda Boi Dourado 3J" in fazendas
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_fazendas_visitadas_global():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            primeira = visitas.iniciar_visita("5500000000001", tecnico_nome="Danilo")
            visitas.atualizar_campo(primeira["id"], "data_visita", "2026-05-10")
            visitas.atualizar_campo(primeira["id"], "fazenda", "Fazenda Imperial")
            segunda = visitas.iniciar_visita("5500000000002", tecnico_nome="Marcelo")
            visitas.atualizar_campo(segunda["id"], "data_visita", "2026-06-17")
            visitas.atualizar_campo(segunda["id"], "fazenda", "Fazenda Boi Dourado 3J")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "fazendas visitadas")

            assert reply is None
            workbook = load_workbook(BytesIO(sent[0][4]))
            sheet = workbook["Visitas"]
            fazendas = [
                row[0]
                for row in sheet.iter_rows(
                    min_row=2,
                    min_col=5,
                    max_col=5,
                    values_only=True,
                )
            ]
            assert "Fazenda Imperial" in fazendas
            assert "Fazenda Boi Dourado 3J" in fazendas
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_localizacao_visita_por_id():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            rdv, visitas, _sender = _install_services(temp_dir)
            other = rdv.get_collaborator_by_phone("5500000000002")["telefone_whatsapp"]
            visita = visitas.iniciar_visita("5500000000001", tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            visitas.adicionar_localizacao(visita["id"], -15.0019124, -50.7714295)

            reply = api_whatsapp.handle_rdv_text_message(
                other,
                f"localizacao visita {visita['id']}",
            )

            assert "Fazenda Imperial" in reply
            assert "https://maps.google.com/?q=-15.0019124,-50.7714295" in reply
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_status_continua_por_telefone():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            rdv, visitas, sender = _install_services(temp_dir)
            other = rdv.get_collaborator_by_phone("5500000000002")["telefone_whatsapp"]
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            reply = api_whatsapp.handle_rdv_text_message(other, "visita status")

            assert reply == api_whatsapp.NO_OPEN_VISITA_MESSAGE
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_canceladas_nao_aparecem():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            cancelada = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(cancelada["id"], "fazenda", "Fazenda Cancelada")
            visitas.cancelar_visita(cancelada["id"])
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            list_reply = api_whatsapp.handle_rdv_text_message(sender, "visitas")
            planilha_reply = api_whatsapp.handle_rdv_text_message(sender, "planilha visitas")
            relatorio_reply = api_whatsapp.handle_rdv_text_message(sender, "relatorio visita")
            fazenda_reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "relatorio fazenda Fazenda Cancelada",
            )

            assert list_reply == api_whatsapp.NO_VALID_VISITA_MESSAGE
            assert planilha_reply is None
            assert relatorio_reply == "Você ainda não possui relatórios de visitas finalizadas."
            assert "Não encontrei visita técnica válida" in fazenda_reply
            workbook = load_workbook(BytesIO(sent[0][4]))
            assert workbook["Visitas"].max_row == 1
            assert len(sent) == 1
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_localizacao_visita_ignora_cancelada():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            visitas.adicionar_localizacao(visita["id"], -15.0019124, -50.7714295)
            visitas.cancelar_visita(visita["id"])

            reply = api_whatsapp.handle_rdv_text_message(sender, "localização visita")

            assert reply == api_whatsapp.NO_VALID_VISITA_MESSAGE
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_status_apos_cancelar():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, _visitas, sender = _install_services(temp_dir)
            api_whatsapp.handle_rdv_text_message(sender, "visita")
            cancel_reply = api_whatsapp.handle_rdv_text_message(sender, "cancelar visita")

            status_reply = api_whatsapp.handle_rdv_text_message(sender, "visita status")

            assert cancel_reply == "Visita cancelada com sucesso."
            assert status_reply == api_whatsapp.NO_OPEN_VISITA_MESSAGE
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_comandos_visita_aceitam_acentos_e_sem_acentos():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender, tecnico_nome="Danilo")
            visitas.atualizar_campo(visita["id"], "fazenda", "Fazenda Imperial")
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            visitas.adicionar_localizacao(visita["id"], -15.0019124, -50.7714295)
            sent = []

            def fake_send(to, content, filename, caption, mime_type):
                sent.append((to, filename, caption, mime_type, content))

            api_whatsapp.send_whatsapp_document = fake_send

            assert api_whatsapp.handle_rdv_text_message(
                sender, "localizacao visita"
            ).startswith("Fazenda Imperial")
            assert api_whatsapp.handle_rdv_text_message(
                sender, "localização visita"
            ).startswith("Fazenda Imperial")
            assert api_whatsapp.handle_rdv_text_message(
                sender,
                f"relatorio visita {visita['id']}",
            ) is None
            assert api_whatsapp.handle_rdv_text_message(
                sender,
                f"relatório visita {visita['id']}",
            ) is None
            assert len(sent) == 2
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_rdv_nao_quebrou_comandos_principais():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_weekly = api_whatsapp._send_weekly_rdv_excel
    original_monthly = api_whatsapp._send_monthly_rdv_excel
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, _, sender = _install_services(temp_dir)
            monthly = []
            weekly = []
            api_whatsapp._send_monthly_rdv_excel = lambda phone, month="": monthly.append(
                (phone, month)
            )
            api_whatsapp._send_weekly_rdv_excel = lambda phone, week="": weekly.append(
                (phone, week)
            )

            assert "Resumo geral do mes" in api_whatsapp.handle_rdv_text_message(sender, "resumo")
            assert api_whatsapp.handle_rdv_text_message(sender, "planilha") is None
            assert "Resumo geral da semana" in api_whatsapp.handle_rdv_text_message(
                sender, "resumo semanal"
            )
            assert api_whatsapp.handle_rdv_text_message(sender, "planilha semanal") is None
            assert "Qual a cidade/local de origem?" in api_whatsapp.handle_rdv_text_message(
                sender, "km inicio 1000"
            )
            assert "Antes de finalizar" in api_whatsapp.handle_rdv_text_message(
                sender, "km termino 1200"
            )
            assert "Viagem cancelada com sucesso." in api_whatsapp.handle_rdv_text_message(
                sender, "km cancelar"
            )
            assert monthly
            assert weekly
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp._send_weekly_rdv_excel = original_weekly
        api_whatsapp._send_monthly_rdv_excel = original_monthly


class _FakeVideoMediaService:
    def __init__(self, *, limit=3, error=None):
        self.limit = limit
        self.error = error
        self.uploads = []

    def video_limit_per_visit(self):
        return self.limit

    def validate_video_file(self, local_path):
        if isinstance(self.error, api_whatsapp.VideoTooLargeError):
            raise self.error
        path = Path(local_path)
        return path.stat().st_size if path.exists() else 5

    def calculate_video_sha256(self, local_path):
        path = Path(local_path)
        content = path.read_bytes() if path.exists() else b""
        import hashlib

        return hashlib.sha256(content).hexdigest()

    def upload_visit_video(self, visita_id, local_path, video_id, mime_type=""):
        if self.error is not None:
            raise self.error
        self.uploads.append(
            {
                "visita_id": visita_id,
                "local_path": Path(local_path),
                "video_id": video_id,
                "mime_type": mime_type,
            }
        )
        return {
            "bucket": "lucre-agro-midias",
            "storage_key": f"visitas/test/{video_id}.mp4",
            "public_url": f"https://cdn.example/{video_id}.mp4",
            "size_bytes": Path(local_path).stat().st_size if Path(local_path).exists() else 5,
            "content_type": mime_type or "video/mp4",
        }


def test_visita_iniciar_fluxo():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            reply = api_whatsapp.handle_rdv_text_message(sender, "visita")

            assert "Vamos iniciar uma visita" in reply
            assert "fazenda ou propriedade" in reply
            assert visitas.obter_visita_aberta(sender)["estado_fluxo"] == "aguardando_fazenda"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_preencher_campos():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            assert "proprietário" in api_whatsapp.handle_rdv_text_message(
                sender, "Fazenda Imperial"
            ).lower()
            assert "telefone" in api_whatsapp.handle_rdv_text_message(
                sender, "Alexander Duarte Paniago"
            ).lower()
            api_whatsapp.handle_rdv_text_message(sender, "(61) 99999-8888")
            api_whatsapp.handle_rdv_text_message(sender, "Paulo Silva")
            api_whatsapp.handle_rdv_text_message(sender, "nao informado")
            tamanho = api_whatsapp.handle_rdv_text_message(sender, "https://maps.google.com/?q=-15,-50")
            descricao = api_whatsapp.handle_rdv_text_message(sender, "500 hectares")
            assert "Qual a área total da fazenda?" in tamanho
            assert "Descrição da visita" in descricao
            assert "Você pode responder digitando o texto ou enviando um áudio." in descricao
            assert "o sistema fará a transcrição automaticamente." in descricao
            obs = api_whatsapp.handle_rdv_text_message(
                sender, "Apresentacao de produtos ao cliente"
            )
            assert "Observações adicionais" in obs
            assert "Você pode responder por texto ou áudio." in obs
            assert "plantações específicas" in obs
            api_whatsapp.handle_rdv_text_message(sender, "Pedido de 300T")
            final = api_whatsapp.handle_rdv_text_message(sender, "finalizar observacoes")

            visita = visitas.obter_visita_aberta(sender)
            assert "Observações salvas" in final
            assert visita["estado_fluxo"] == "visita_aberta"
            assert visita["fazenda"] == "Fazenda Imperial"
            assert visita["proprietario"] == "Alexander Duarte Paniago"
            assert visita["telefone_proprietario"] == "61999998888"
            assert visita["gerente"] == "Paulo Silva"
            assert visita["telefone_gerente"] == "Não informado"
            assert visita["localizacao_texto"] == "https://maps.google.com/?q=-15,-50"
            assert visita["maps_url_principal"] == "https://maps.google.com/?q=-15,-50"
            assert visita["area"] == "500 hectares"
            assert visita["descricao_visita"] == "Apresentacao de produtos ao cliente"
            assert "Pedido de 300T" in visita["observacoes_gerais"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_pergunta_area_total_nao_usa_texto_antigo():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            api_whatsapp.handle_rdv_text_message(sender, "Fazenda Imperial")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "https://maps.google.com/?q=-15,-50",
            )

            assert "Qual a área total da fazenda?" in reply
            assert "Qual área, talhão ou local da propriedade foi visitado?" not in reply
            assert "área, talhão ou local" not in reply.lower()
            assert "Área/local visitado" not in reply
            assert visitas.obter_visita_aberta(sender)["estado_fluxo"] == "aguardando_area"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_area_total_pode_ser_ignorada_com_aliases():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        for skip_text in ("pular", "não sei", "sem informação"):
            with tempfile.TemporaryDirectory() as temp_dir:
                _, visitas, sender = _install_services(temp_dir)
                visita = visitas.iniciar_visita(sender)
                visitas.atualizar_campo(visita["id"], "estado_fluxo", "aguardando_area")

                reply = api_whatsapp.handle_rdv_text_message(sender, skip_text)
                saved = visitas.obter_visita_aberta(sender)

                assert "Descrição da visita" in reply
                assert saved["estado_fluxo"] == "aguardando_descricao_visita"
                assert not saved["area"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_valida_telefone_invalido_e_pede_correcao():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            api_whatsapp.handle_rdv_text_message(sender, "Fazenda Imperial")
            api_whatsapp.handle_rdv_text_message(sender, "João da Silva")
            reply = api_whatsapp.handle_rdv_text_message(sender, "12345")

            visita = visitas.obter_visita_aberta(sender)
            assert "Telefone inválido" in reply
            assert "62999998888" in reply
            assert visita["estado_fluxo"] == "aguardando_telefone_proprietario"
            assert not visita["telefone_proprietario"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_descricao_curta_pede_mais_detalhes():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            api_whatsapp.handle_rdv_text_message(sender, "Fazenda Imperial")
            api_whatsapp.handle_rdv_text_message(sender, "João da Silva")
            api_whatsapp.handle_rdv_text_message(sender, "61999998888")
            api_whatsapp.handle_rdv_text_message(sender, "Paulo Silva")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "500 hectares")
            reply = api_whatsapp.handle_rdv_text_message(sender, "curta")

            visita = visitas.obter_visita_aberta(sender)
            assert "descrição ficou muito curta" in reply
            assert visita["estado_fluxo"] == "aguardando_descricao_visita"
            assert not visita["descricao_visita"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_pular_funciona_em_campos_opcionais():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            api_whatsapp.handle_rdv_text_message(sender, "Fazenda Imperial")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")

            visita = visitas.obter_visita_aberta(sender)
            assert visita["proprietario"] == "Não informado"
            assert visita["telefone_proprietario"] == "Não informado"
            assert visita["gerente"] == "Não informado"
            assert visita["telefone_gerente"] == "Não informado"
            assert visita["estado_fluxo"] == "aguardando_localizacao"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_pede_localizacao_explicitamente_e_aceita_endereco():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            api_whatsapp.handle_rdv_text_message(sender, "Fazenda Imperial")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            api_whatsapp.handle_rdv_text_message(sender, "pular")
            reply = api_whatsapp.handle_rdv_text_message(sender, "pular")
            next_reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "Estrada municipal, 10 km depois da ponte",
            )

            visita = visitas.obter_visita_aberta(sender)
            assert "localização da fazenda/propriedade" in reply
            assert "Compartilhar a localização pelo WhatsApp" in reply
            assert "Qual a área total da fazenda?" in next_reply
            assert visita["localizacao_texto"] == "Estrada municipal, 10 km depois da ponte"
            assert not visita["maps_url_principal"]
            assert visita["estado_fluxo"] == "aguardando_area"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_localizacao_pode_ser_ignorada_com_aliases():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        for skip_text in ("pular", "não sei", "sem informação"):
            with tempfile.TemporaryDirectory() as temp_dir:
                _, visitas, sender = _install_services(temp_dir)

                api_whatsapp.handle_rdv_text_message(sender, "visita")
                api_whatsapp.handle_rdv_text_message(sender, "Fazenda Imperial")
                api_whatsapp.handle_rdv_text_message(sender, "pular")
                api_whatsapp.handle_rdv_text_message(sender, "pular")
                api_whatsapp.handle_rdv_text_message(sender, "pular")
                api_whatsapp.handle_rdv_text_message(sender, "pular")
                reply = api_whatsapp.handle_rdv_text_message(sender, skip_text)

                visita = visitas.obter_visita_aberta(sender)
                assert "Qual a área total da fazenda?" in reply
                assert not visita["localizacao_texto"]
                assert visita["estado_fluxo"] == "aguardando_area"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_localizacao_nativa_durante_etapa_salva_e_avanca():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "aguardando_localizacao")

            reply = api_whatsapp.handle_visitas_location_message(
                sender,
                {"latitude": -15.0019124, "longitude": -50.7714295},
            )

            saved = visitas.obter_visita_completa(visita["id"])
            assert "Localização salva" in reply
            assert "Qual a área total da fazenda?" in reply
            assert saved["estado_fluxo"] == "aguardando_area"
            assert saved["maps_url_principal"] == (
                "https://maps.google.com/?q=-15.0019124,-50.7714295"
            )
            assert len(saved["localizacoes"]) == 1
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_aceita_nomes_validos_com_acentos():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            api_whatsapp.handle_rdv_text_message(sender, "Fazenda São José")
            api_whatsapp.handle_rdv_text_message(sender, "José Antônio")

            visita = visitas.obter_visita_aberta(sender)
            assert visita["fazenda"] == "Fazenda São José"
            assert visita["proprietario"] == "José Antônio"
            assert visita["estado_fluxo"] == "aguardando_telefone_proprietario"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_rejeita_texto_claramente_invalido():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            reply = api_whatsapp.handle_rdv_text_message(sender, "jfadlojasd")

            visita = visitas.obter_visita_aberta(sender)
            assert "Não consegui entender" in reply
            assert visita["estado_fluxo"] == "aguardando_fazenda"
            assert not visita["fazenda"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_comando_global_cancelar_continua_funcionando():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            reply = api_whatsapp.handle_rdv_text_message(sender, "cancelar")

            assert reply == "Visita cancelada com sucesso."
            assert visitas.obter_visita_aberta(sender) is None
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_observacoes_gerais_multiplas():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "aguardando_observacoes_gerais")

            api_whatsapp.handle_rdv_text_message(sender, "Primeira observacao")
            api_whatsapp.handle_rdv_text_message(sender, "Segunda observacao")
            reply = api_whatsapp.handle_rdv_text_message(sender, "pronto")

            saved = visitas.obter_visita_aberta(sender)
            assert "Observações salvas" in reply
            assert saved["estado_fluxo"] == "visita_aberta"
            assert saved["observacoes_gerais"].splitlines() == [
                "Primeira observacao",
                "Segunda observacao",
            ]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_observacao_digitada_curta_continua_funcionando():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(
                visita["id"], "estado_fluxo", "aguardando_observacoes_gerais"
            )

            reply = api_whatsapp.handle_rdv_text_message(
                sender, "Aplicação acompanhada sem intercorrências."
            )

            saved = visitas.obter_visita_aberta(sender)
            assert "Observação salva" in reply
            assert saved["observacoes_gerais"] == (
                "Aplicação acompanhada sem intercorrências."
            )
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_observacao_de_audio_acima_de_1000_caracteres_e_aceita():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(
                visita["id"], "estado_fluxo", "aguardando_observacoes_gerais"
            )
            transcription = ("manejo observado durante a visita " * 45).strip()

            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                transcription,
                is_audio_transcription=True,
            )

            saved = visitas.obter_visita_aberta(sender)
            assert len(transcription) > 1000
            assert "Observação salva" in reply
            assert saved["observacoes_gerais"] == transcription
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_observacao_longa_de_audio_e_dividida_em_ordem(monkeypatch):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(
                visita["id"], "estado_fluxo", "aguardando_observacoes_gerais"
            )
            monkeypatch.setenv("VISITA_OBSERVACAO_MAX_CHARS", "100")
            monkeypatch.setenv("VISITA_OBSERVACAO_TOTAL_MAX_CHARS", "500")
            transcription = " ".join(f"item-{index:02d}" for index in range(40))

            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                transcription,
                is_audio_transcription=True,
            )

            saved = visitas.obter_visita_aberta(sender)
            parts = saved["observacoes_gerais"].splitlines()
            assert reply == (
                f"Áudio transcrito e salvo em {len(parts)} observações do relatório."
            )
            assert len(parts) > 1
            assert all(len(part) <= 100 for part in parts)
            assert " ".join(parts) == transcription
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_observacao_de_audio_acima_do_teto_total_e_rejeitada(monkeypatch):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(
                visita["id"], "estado_fluxo", "aguardando_observacoes_gerais"
            )
            monkeypatch.setenv("VISITA_OBSERVACAO_MAX_CHARS", "100")
            monkeypatch.setenv("VISITA_OBSERVACAO_TOTAL_MAX_CHARS", "200")

            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "texto de campo " * 30,
                is_audio_transcription=True,
            )

            saved = visitas.obter_visita_aberta(sender)
            assert "dividido em partes menores" in reply
            assert not saved["observacoes_gerais"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_descricao_aceita_mais_de_1000_caracteres():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(
                visita["id"], "estado_fluxo", "aguardando_descricao_visita"
            )
            description = ("Descrição técnica detalhada da lavoura. " * 35).strip()

            reply = api_whatsapp.handle_rdv_text_message(sender, description)

            saved = visitas.obter_visita_aberta(sender)
            assert len(description) > 1000
            assert "Observações adicionais" in reply
            assert saved["descricao_visita"] == description
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_descricao_acima_do_limite_configurado_tem_erro_amigavel(
    monkeypatch,
):
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(
                visita["id"], "estado_fluxo", "aguardando_descricao_visita"
            )
            monkeypatch.setenv("VISITA_DESCRICAO_MAX_CHARS", "100")

            reply = api_whatsapp.handle_rdv_text_message(
                sender, "descrição detalhada " * 10
            )

            saved = visitas.obter_visita_aberta(sender)
            assert reply == (
                "A descrição ficou muito longa. "
                "Envie um resumo menor ou divida em partes."
            )
            assert not saved["descricao_visita"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_campo_curto_continua_rejeitando_texto_absurdo():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)

            api_whatsapp.handle_rdv_text_message(sender, "visita")
            reply = api_whatsapp.handle_rdv_text_message(sender, "F" * 121)

            saved = visitas.obter_visita_aberta(sender)
            assert "no máximo 120 caracteres" in reply
            assert saved["estado_fluxo"] == "aguardando_fazenda"
            assert not saved["fazenda"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas


def test_visita_foto_com_comentario_individual():
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            reply = api_whatsapp.handle_visitas_media_message(
                sender, "image", "wamid.1", str(Path(temp_dir) / "foto1.jpg")
            )
            assert "Foto 1 adicionada" in reply
            assert "1 - Sim" in reply
            assert "Digite o comentario da Foto 1" in api_whatsapp.handle_rdv_text_message(sender, "1")
            done = api_whatsapp.handle_rdv_text_message(sender, "Vazamento no registro")

            saved = visitas.obter_visita_completa(visita["id"])
            assert "Fotos salvas" in done
            assert saved["midias"][0]["comentario"] == "Vazamento no registro"
            assert saved["midias"][0]["comentario_status"] == "resolvido"
    finally:
        api_whatsapp.visitas_service = original_visitas


def test_visita_duas_fotos_em_fila_e_bloqueia_fechamento():
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            api_whatsapp.handle_visitas_media_message(sender, "image", "wamid.1", str(Path(temp_dir) / "foto1.jpg"))
            api_whatsapp.handle_visitas_media_message(sender, "image", "wamid.2", str(Path(temp_dir) / "foto2.jpg"))
            blocked = api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            assert "Antes de fechar" in blocked

            next_reply = api_whatsapp.handle_rdv_text_message(sender, "2")
            assert "Foto 2" in next_reply
            done = api_whatsapp.handle_rdv_text_message(sender, "sem comentario")
            saved = visitas.obter_visita_completa(visita["id"])
            assert "Fotos salvas" in done
            assert [media["comentario"] for media in saved["midias"]] == [
                "Sem comentario informado.",
                "Sem comentario informado.",
            ]
    finally:
        api_whatsapp.visitas_service = original_visitas


def test_visita_duas_fotos_comentario_direto_na_primeira_e_pula_segunda():
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            first = api_whatsapp.handle_visitas_media_message(
                sender, "image", "wamid.1", str(Path(temp_dir) / "foto1.jpg")
            )
            second = api_whatsapp.handle_visitas_media_message(
                sender, "image", "wamid.2", str(Path(temp_dir) / "foto2.jpg")
            )
            next_reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "Vazamento no registro da Foto 1",
            )
            done = api_whatsapp.handle_rdv_text_message(sender, "pular")

            saved = visitas.obter_visita_completa(visita["id"])
            assert "Foto 1 recebida" in first
            assert "Foto 2 recebida" in second
            assert "comentario da Foto 1" in second
            assert "Foto 2" in next_reply
            assert "Fotos salvas" in done
            assert [media["comentario"] for media in saved["midias"]] == [
                "Vazamento no registro da Foto 1",
                "Sem comentario informado.",
            ]
            assert [media["comentario_status"] for media in saved["midias"]] == [
                "resolvido",
                "resolvido",
            ]
    finally:
        api_whatsapp.visitas_service = original_visitas


def test_visita_fotos_em_fila_nao_disparam_fallback_generico():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_download = api_whatsapp.download_media
    original_send_text = api_whatsapp.send_whatsapp_text
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            sent = []

            def fake_download(media_id, destino):
                Path(destino).write_bytes(b"foto")
                return Path(destino)

            api_whatsapp.download_media = fake_download
            api_whatsapp.send_whatsapp_text = lambda to, message: sent.append((to, message))

            api_whatsapp._handle_whatsapp_message(
                {
                    "from": sender,
                    "id": "wamid.foto.1",
                    "type": "image",
                    "timestamp": "1780000000",
                    "image": {"id": "foto-1", "mime_type": "image/jpeg"},
                }
            )
            api_whatsapp._handle_whatsapp_message(
                {
                    "from": sender,
                    "id": "wamid.foto.2",
                    "type": "image",
                    "timestamp": "1780000001",
                    "image": {"id": "foto-2", "mime_type": "image/jpeg"},
                }
            )

            replies = [message for _, message in sent]
            assert len(visitas.obter_visita_completa(visita["id"])["midias"]) == 2
            assert all(
                "Recebi sua mensagem, mas por enquanto consigo processar apenas imagem ou documento."
                not in reply
                for reply in replies
            )
            assert any("Foto 2 recebida" in reply for reply in replies)
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.download_media = original_download
        api_whatsapp.send_whatsapp_text = original_send_text
        api_whatsapp.visita_active_states.clear()


def test_visita_tres_fotos_mantem_fila_de_comentarios():
    original_visitas = api_whatsapp.visitas_service
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            first = api_whatsapp.handle_visitas_media_message(
                sender, "image", "wamid.1", str(Path(temp_dir) / "foto1.jpg")
            )
            second = api_whatsapp.handle_visitas_media_message(
                sender, "image", "wamid.2", str(Path(temp_dir) / "foto2.jpg")
            )
            third = api_whatsapp.handle_visitas_media_message(
                sender, "image", "wamid.3", str(Path(temp_dir) / "foto3.jpg")
            )

            assert "Foto 1 adicionada" in first
            assert "Foto 1 adicionada" in second
            assert "Foto 1 adicionada" in third
            assert "Digite o comentario da Foto 1" in api_whatsapp.handle_rdv_text_message(sender, "1")
            assert "Foto 2" in api_whatsapp.handle_rdv_text_message(sender, "Comentario da foto 1")
            assert "Digite o comentario da Foto 2" in api_whatsapp.handle_rdv_text_message(sender, "sim")
            assert "Foto 3" in api_whatsapp.handle_rdv_text_message(sender, "Comentario da foto 2")
            assert "Fotos salvas" in api_whatsapp.handle_rdv_text_message(sender, "2")

            saved = visitas.obter_visita_completa(visita["id"])
            assert [media["comentario"] for media in saved["midias"]] == [
                "Comentario da foto 1",
                "Comentario da foto 2",
                "Sem comentario informado.",
            ]
    finally:
        api_whatsapp.visitas_service = original_visitas


def test_video_sem_visita_ativa_responde_orientacao():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_send_text = api_whatsapp.send_whatsapp_text
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            sent = []
            downloads = []
            api_whatsapp.send_whatsapp_text = lambda to, message: sent.append(
                (to, message)
            )
            api_whatsapp.download_media = lambda media_id, destino: downloads.append(
                media_id
            )

            api_whatsapp._handle_whatsapp_message(
                {
                    "from": sender,
                    "id": "wamid.video.sem.visita",
                    "type": "video",
                    "timestamp": "1780000000",
                    "video": {
                        "id": "video-sem-visita",
                        "mime_type": "video/mp4",
                    },
                }
            )

            assert sent == [(sender, api_whatsapp.VISITA_VIDEO_NO_OPEN_MESSAGE)]
            assert downloads == []
            assert visitas.listar_visitas_validas()["visitas"] == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_text = original_send_text
        api_whatsapp.download_media = original_download


def test_video_durante_visita_upload_salva_metadados_e_pede_legenda():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            downloaded = Path(temp_dir) / "video.mp4"

            def fake_download(media_id, destino):
                Path(destino).write_bytes(b"video")
                return Path(destino)

            api_whatsapp.download_media = fake_download
            api_whatsapp.visita_media_service = _FakeVideoMediaService()

            reply = api_whatsapp.handle_visitas_video_message(
                sender,
                "video-1",
                "video/mp4",
            )

            saved = visitas.obter_visita_completa(visita["id"])
            video = saved["midias"][0]
            assert "Vídeo recebido e anexado" in reply
            assert "legenda" in reply
            assert video["tipo"] == "video"
            assert video["storage_key"] == "visitas/test/video-1.mp4"
            assert video["public_url"] == "https://cdn.example/video-1.mp4"
            assert video["tamanho_bytes"] == 5
            assert video["mime_type"] == "video/mp4"
            assert video["video_hash"] == (
                "0cab1c9617404faf2b24e221e189ca5945813e14"
                "d3f766345b09ca13bbe28ffc"
            )
            assert video["comentario_status"] == "pendente"
            assert not downloaded.exists()
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_video_durante_revisao_anexa_salva_legenda_e_nao_finaliza():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            api_whatsapp.send_whatsapp_document = lambda *args, **kwargs: None

            def fake_download(media_id, destino):
                Path(destino).write_bytes(b"video")
                return Path(destino)

            api_whatsapp.download_media = fake_download
            api_whatsapp.visita_media_service = _FakeVideoMediaService()
            api_whatsapp.handle_rdv_text_message(sender, "fechar visita")

            reply = api_whatsapp.handle_visitas_video_message(sender, "video-review", "video/mp4")
            legend = api_whatsapp.handle_rdv_text_message(sender, "Video complementar")

            saved = visitas.obter_visita_completa(visita["id"])
            video = saved["midias"][0]
            assert "Vídeo recebido e anexado" in reply
            assert "prévia anterior ficará desatualizada" in reply.lower()
            assert "Legenda salva" in legend
            assert "prévia anterior pode estar desatualizada" in legend.lower()
            assert video["tipo"] == "video"
            assert video["comentario"] == "Video complementar"
            assert video["public_url"] == "https://cdn.example/video-review.mp4"
            assert saved["status"] == "aberta"
            assert saved["estado_fluxo"] == "aguardando_revisao_final"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.send_whatsapp_document = original_sender
        api_whatsapp.visita_active_states.clear()


def test_dois_videos_em_sequencia_mantem_fila_de_legendas():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            def fake_download(media_id, destino):
                Path(destino).write_bytes(media_id.encode("utf-8"))
                return Path(destino)

            api_whatsapp.download_media = fake_download
            api_whatsapp.visita_media_service = _FakeVideoMediaService()

            first = api_whatsapp.handle_visitas_video_message(sender, "video-1", "video/mp4")
            second = api_whatsapp.handle_visitas_video_message(sender, "video-2", "video/mp4")
            next_reply = api_whatsapp.handle_rdv_text_message(sender, "Legenda do video 1")
            done = api_whatsapp.handle_rdv_text_message(sender, "pular")

            saved = visitas.obter_visita_completa(visita["id"])
            assert "Vídeo recebido e anexado" in first
            assert "Vídeo recebido e anexado" in second
            assert "Vídeo 1" in second
            assert "Vídeo 2" in next_reply
            assert "Fotos salvas" in done
            assert [media["comentario"] for media in saved["midias"]] == [
                "Legenda do video 1",
                "Sem comentario informado.",
            ]
            assert [media["comentario_status"] for media in saved["midias"]] == [
                "resolvido",
                "resolvido",
            ]
            assert visitas.obter_visita(visita["id"])["estado_fluxo"] == "visita_aberta"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_video_acima_do_limite_responde_mensagem_amigavel():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            api_whatsapp.download_media = lambda media_id, destino: Path(destino)
            api_whatsapp.visita_media_service = _FakeVideoMediaService(
                error=api_whatsapp.VideoTooLargeError("too_large")
            )

            reply = api_whatsapp.handle_visitas_video_message(sender, "video-big", "video/mp4")

            assert "vídeo ficou muito grande" in reply
            assert "até 15 MB" in reply
            assert visitas.obter_visita_completa(visita["id"])["midias"] == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_video_limite_por_visita_bloqueia_apos_hash_e_nao_faz_upload():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            visitas.adicionar_midia(
                visita["id"],
                "video",
                storage_key="v1",
                video_hash="hash-video-existente",
            )
            downloads = []

            def fake_download(media_id, destino):
                downloads.append(media_id)
                Path(destino).write_bytes(b"novo-video")
                return Path(destino)

            media_service = _FakeVideoMediaService(limit=1)
            api_whatsapp.download_media = fake_download
            api_whatsapp.visita_media_service = media_service

            reply = api_whatsapp.handle_visitas_video_message(sender, "video-2", "video/mp4")

            assert "Limite atual: 1" in reply
            assert downloads == ["video-2"]
            assert media_service.uploads == []
            assert len(visitas.obter_visita_completa(visita["id"])["midias"]) == 1
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_visita_permite_mais_de_tres_videos_ate_limite_configurado():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            def fake_download(media_id, destino):
                Path(destino).write_bytes(media_id.encode("utf-8"))
                return Path(destino)

            media_service = _FakeVideoMediaService(limit=10)
            api_whatsapp.download_media = fake_download
            api_whatsapp.visita_media_service = media_service

            replies = [
                api_whatsapp.handle_visitas_video_message(sender, f"video-{index}", "video/mp4")
                for index in range(1, 5)
            ]

            saved = visitas.obter_visita_completa(visita["id"])
            assert all("anexado" in reply for reply in replies)
            assert saved["contadores"]["videos"] == 4
            assert len(media_service.uploads) == 4
            assert all(media.get("video_hash") for media in saved["midias"])
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_video_duplicado_por_sha256_nao_faz_upload_nem_conta_e_remove_temporario():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            temp_paths = []

            def fake_download(media_id, destino):
                path = Path(destino)
                temp_paths.append(path)
                path.write_bytes(b"mesmo-conteudo")
                return path

            media_service = _FakeVideoMediaService(limit=10)
            api_whatsapp.download_media = fake_download
            api_whatsapp.visita_media_service = media_service

            first = api_whatsapp.handle_visitas_video_message(sender, "video-1", "video/mp4")
            second = api_whatsapp.handle_visitas_video_message(sender, "video-2", "video/mp4")

            saved = visitas.obter_visita_completa(visita["id"])
            assert "anexado" in first
            assert "parece" in second
            assert saved["contadores"]["videos"] == 1
            assert len(saved["midias"]) == 1
            assert len(media_service.uploads) == 1
            assert all(not path.exists() for path in temp_paths)
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_video_hash_antigo_vazio_nao_bloqueia_novo_upload():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            visitas.adicionar_midia(visita["id"], "video", storage_key="video-legado")

            def fake_download(media_id, destino):
                Path(destino).write_bytes(b"conteudo-legado-sem-hash")
                return Path(destino)

            media_service = _FakeVideoMediaService(limit=10)
            api_whatsapp.download_media = fake_download
            api_whatsapp.visita_media_service = media_service

            reply = api_whatsapp.handle_visitas_video_message(sender, "video-novo", "video/mp4")

            saved = visitas.obter_visita_completa(visita["id"])
            assert "anexado" in reply
            assert saved["contadores"]["videos"] == 2
            assert len(media_service.uploads) == 1
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_videos_diferentes_com_nomes_parecidos_sao_permitidos():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")

            def fake_download(media_id, destino):
                content = b"conteudo-a" if media_id == "video-final" else b"conteudo-b"
                Path(destino).write_bytes(content)
                return Path(destino)

            media_service = _FakeVideoMediaService(limit=10)
            api_whatsapp.download_media = fake_download
            api_whatsapp.visita_media_service = media_service

            first = api_whatsapp.handle_visitas_video_message(sender, "video-final", "video/mp4")
            second = api_whatsapp.handle_visitas_video_message(sender, "video_final", "video/mp4")

            saved = visitas.obter_visita_completa(visita["id"])
            assert "anexado" in first
            assert "anexado" in second
            assert saved["contadores"]["videos"] == 2
            assert len({media["video_hash"] for media in saved["midias"]}) == 2
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_video_falha_upload_responde_mensagem_amigavel():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            api_whatsapp.download_media = lambda media_id, destino: Path(destino)
            api_whatsapp.visita_media_service = _FakeVideoMediaService(
                error=api_whatsapp.VideoUploadError("spaces indisponivel")
            )

            reply = api_whatsapp.handle_visitas_video_message(sender, "video-fail", "video/mp4")

            assert reply == api_whatsapp.VISITA_VIDEO_UPLOAD_ERROR_MESSAGE
            assert visitas.obter_visita_completa(visita["id"])["midias"] == []
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_legenda_apos_video_salva_comentario():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_media_service = api_whatsapp.visita_media_service
    original_download = api_whatsapp.download_media
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            api_whatsapp.download_media = lambda media_id, destino: Path(destino)
            api_whatsapp.visita_media_service = _FakeVideoMediaService()
            api_whatsapp.handle_visitas_video_message(sender, "video-1", "video/mp4")

            reply = api_whatsapp.handle_rdv_text_message(
                sender,
                "Área com falha perto da entrada da fazenda",
            )

            saved = visitas.obter_visita_completa(visita["id"])
            assert "Legenda salva" in reply
            assert saved["midias"][0]["comentario"] == (
                "Área com falha perto da entrada da fazenda"
            )
            assert saved["midias"][0]["comentario_status"] == "resolvido"
            assert visitas.obter_visita(visita["id"])["estado_fluxo"] == "visita_aberta"
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.visita_media_service = original_media_service
        api_whatsapp.download_media = original_download
        api_whatsapp.visita_active_states.clear()


def test_visita_resumo_final_mostra_comentarios_das_fotos():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )
            media = visitas.adicionar_midia(
                visita["id"],
                "foto",
                media_id_whatsapp="wamid.1",
                caminho_arquivo=str(Path(temp_dir) / "foto1.jpg"),
            )
            visitas.salvar_comentario_foto(media["id"], "Vazamento no registro")

            reply = api_whatsapp.handle_rdv_text_message(sender, "fechar visita")

            assert "Prévia do relatório enviada" in reply
            assert sent[0][4].startswith(b"%PDF")
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender


def test_visita_fechar():
    original_rdv = api_whatsapp.rdv_service
    original_visitas = api_whatsapp.visitas_service
    original_sender = api_whatsapp.send_whatsapp_document
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            _, visitas, sender = _install_services(temp_dir)
            visita = visitas.iniciar_visita(sender)
            visitas.atualizar_campo(visita["id"], "estado_fluxo", "visita_aberta")
            sent = []
            api_whatsapp.send_whatsapp_document = (
                lambda to, content, filename, caption, mime_type: sent.append(
                    (to, filename, caption, mime_type, content)
                )
            )

            reply = api_whatsapp.handle_rdv_text_message(sender, "fechar visita")
            saved = visitas.obter_visita(visita["id"])

            assert "Prévia do relatório enviada" in reply
            assert "Revise os dados antes de finalizar" in reply
            assert "Tipo" not in reply
            assert "1. Finalizar visita" in reply
            assert sent[0][4].startswith(b"%PDF")
            assert saved["status"] == "aberta"
            assert saved["estado_fluxo"] == "aguardando_revisao_final"

            final = api_whatsapp.handle_rdv_text_message(sender, "1")
            closed = visitas.obter_visita(visita["id"])
            assert "Visita finalizada com sucesso." in final
            assert closed["status"] == "fechada"
            assert closed["fechado_em"]
    finally:
        api_whatsapp.rdv_service = original_rdv
        api_whatsapp.visitas_service = original_visitas
        api_whatsapp.send_whatsapp_document = original_sender
