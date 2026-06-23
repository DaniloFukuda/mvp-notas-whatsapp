from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DATE_FORMAT = "dd/mm/yyyy"
DATETIME_FORMAT = "dd/mm/yyyy hh:mm"


def build_visitas_workbook(visitas_data: dict) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    visitas = visitas_data.get("visitas") or []
    midias = visitas_data.get("midias") or []
    localizacoes = visitas_data.get("localizacoes") or []
    dados = visitas_data.get("dados_coletados") or []

    by_visit = {int(visita["id"]): visita for visita in visitas}

    _build_visitas_sheet(workbook, visitas)
    _build_fotos_sheet(workbook, midias, by_visit)
    _build_localizacoes_sheet(workbook, localizacoes, by_visit)
    _build_dados_sheet(workbook, dados, by_visit)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_visitas_sheet(workbook: Workbook, visitas: list[dict]) -> None:
    sheet = workbook.create_sheet("Visitas")
    sheet.append(
        (
            "ID",
            "Data",
            "Técnico",
            "Telefone",
            "Fazenda",
            "Proprietário",
            "Gerente",
            "Área ha",
            "Área alqueires",
            "Safra",
            "Tipo de visita",
            "Status",
            "Qtd fotos",
            "Qtd localizações",
            "Link GPS principal",
            "Observações",
        )
    )
    for visita in visitas:
        sheet.append(
            (
                visita.get("id"),
                _parse_date(visita.get("data_visita")),
                visita.get("tecnico_nome") or "",
                visita.get("telefone_origem") or "",
                visita.get("fazenda") or "",
                visita.get("proprietario") or "",
                visita.get("gerente") or "",
                _optional_float(visita.get("area_hectares")),
                _optional_float(visita.get("area_alqueires")),
                visita.get("safra") or "",
                visita.get("tipo_visita") or "",
                _humanize(visita.get("status")),
                len(visita.get("midias") or []),
                len(visita.get("localizacoes") or []),
                visita.get("maps_url_principal") or "",
                visita.get("observacoes") or "",
            )
        )
    _prepare_sheet(sheet)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 2).number_format = DATE_FORMAT


def _build_visitas_sheet(workbook: Workbook, visitas: list[dict]) -> None:
    sheet = workbook.create_sheet("Visitas")
    sheet.append(
        (
            "ID",
            "Data",
            "Tecnico",
            "Telefone",
            "Fazenda",
            "Proprietario",
            "Telefone do proprietario",
            "Gerente",
            "Telefone do gerente",
            "Area/local visitado",
            "Descricao da visita",
            "Status",
            "Qtd fotos",
            "Qtd localizacoes",
            "Link GPS principal",
            "Observacoes gerais",
        )
    )
    for visita in visitas:
        sheet.append(
            (
                visita.get("id"),
                _parse_date(visita.get("data_visita")),
                visita.get("tecnico_nome") or "",
                visita.get("telefone_origem") or "",
                visita.get("fazenda") or "",
                visita.get("proprietario") or "",
                visita.get("telefone_proprietario") or "",
                visita.get("gerente") or "",
                visita.get("telefone_gerente") or "",
                visita.get("area") or "",
                visita.get("descricao_visita") or visita.get("tipo_visita") or "",
                _humanize(visita.get("status")),
                len(visita.get("midias") or []),
                len(visita.get("localizacoes") or []),
                visita.get("maps_url_principal") or "",
                visita.get("observacoes_gerais") or visita.get("observacoes") or "",
            )
        )
    _prepare_sheet(sheet)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 2).number_format = DATE_FORMAT


def _build_localizacoes_sheet(
    workbook: Workbook,
    rows: list[dict],
    by_visit: dict[int, dict],
) -> None:
    sheet = workbook.create_sheet("Localizações")
    sheet.append(
        (
            "ID visita",
            "Fazenda",
            "Data/hora",
            "Descrição",
            "Latitude",
            "Longitude",
            "Link GPS",
        )
    )
    for row in rows:
        visita = by_visit.get(int(row.get("visita_id") or 0), {})
        sheet.append(
            (
                row.get("visita_id"),
                visita.get("fazenda") or "",
                _parse_datetime(row.get("enviado_em")),
                row.get("descricao") or "",
                _optional_float(row.get("latitude")),
                _optional_float(row.get("longitude")),
                row.get("maps_url") or "",
            )
        )
    _prepare_sheet(sheet)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 3).number_format = DATETIME_FORMAT


def _build_fotos_sheet(
    workbook: Workbook,
    rows: list[dict],
    by_visit: dict[int, dict],
) -> None:
    sheet = workbook.create_sheet("Fotos")
    sheet.append(
        (
            "ID visita",
            "Fazenda",
            "Data/hora",
            "Tipo",
            "Comentario",
            "Latitude",
            "Longitude",
            "Link GPS",
            "Arquivo",
        )
    )
    for row in rows:
        visita = by_visit.get(int(row.get("visita_id") or 0), {})
        sheet.append(
            (
                row.get("visita_id"),
                visita.get("fazenda") or "",
                _parse_datetime(row.get("enviado_em")),
                row.get("tipo") or "",
                row.get("comentario") or row.get("legenda") or "",
                _optional_float(row.get("latitude")),
                _optional_float(row.get("longitude")),
                row.get("maps_url") or "",
                _safe_file_reference(row.get("caminho_arquivo")),
            )
        )
    _prepare_sheet(sheet)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 3).number_format = DATETIME_FORMAT


def _build_dados_sheet(
    workbook: Workbook,
    rows: list[dict],
    by_visit: dict[int, dict],
) -> None:
    sheet = workbook.create_sheet("Dados coletados")
    sheet.append(
        (
            "ID visita",
            "Fazenda",
            "Data/hora",
            "Chave",
            "Valor",
            "Observação",
        )
    )
    for row in rows:
        visita = by_visit.get(int(row.get("visita_id") or 0), {})
        sheet.append(
            (
                row.get("visita_id"),
                visita.get("fazenda") or "",
                _parse_datetime(row.get("criado_em")),
                row.get("chave") or "",
                row.get("valor") or "",
                row.get("observacao") or "",
            )
        )
    _prepare_sheet(sheet)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 3).number_format = DATETIME_FORMAT


def _prepare_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_cells in sheet.columns:
        max_length = max(len(_display_value(cell.value)) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _parse_date(value: object):
    text = str(value or "").strip()
    if not text:
        return None
    for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return text


def _parse_datetime(value: object):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value or "").strip()
    if not text:
        return None
    for date_format in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return text


def _optional_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def _safe_file_reference(value: object) -> str:
    text = str(value or "").strip()
    return Path(text).name if text else ""


def _humanize(value: object) -> str:
    return str(value or "").replace("_", " ").title()


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    return str(value)
