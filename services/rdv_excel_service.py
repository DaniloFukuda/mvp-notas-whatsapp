from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CURRENCY_FORMAT = 'R$ #,##0.00'
DATE_FORMAT = "dd/mm/yyyy"
DATETIME_FORMAT = "dd/mm/yyyy hh:mm"
DISTANCE_FORMAT = '0.00 "km"'


def build_weekly_rdv_workbook(report_data: dict) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    launches = report_data.get("lancamentos") or []
    collaborators = report_data.get("resumo_colaboradores") or []
    categories = report_data.get("resumo_categorias") or []
    pending = report_data.get("pendencias") or []

    _build_launches_sheet(workbook, launches, "Lancamentos")
    _build_collaborators_sheet(workbook, collaborators)
    _build_categories_sheet(workbook, categories)
    _build_launches_sheet(workbook, pending, "Pendencias")

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_launches_sheet(workbook: Workbook, rows: list[dict], title: str) -> None:
    sheet = workbook.create_sheet(title)
    headers = (
        "Data",
        "Colaborador",
        "Telefone",
        "Categoria",
        "Valor",
        "Fornecedor detectado",
        "Data detectada",
        "Origem do valor",
        "Chave de acesso / QR Code / URL",
        "Cidade origem",
        "Cidade destino",
        "KM inicial",
        "KM final",
        "KM rodado",
        "Status",
        "Observacao",
        "Tipo de entrada",
        "Comprovante/arquivo",
        "Recebido em",
    )
    sheet.append(headers)
    for row in rows:
        sheet.append(
            (
                _parse_date(row.get("data_despesa")),
                row.get("colaborador") or "",
                row.get("telefone_origem") or "",
                _humanize(row.get("categoria")),
                float(row.get("valor") or 0),
                row.get("fornecedor_detectado") or row.get("fornecedor") or "",
                _parse_date(row.get("data_detectada")),
                _humanize(row.get("origem_valor")),
                _fiscal_reference(row),
                row.get("cidade_origem") or "",
                row.get("cidade_destino") or "",
                _optional_float(row.get("km_inicio")),
                _optional_float(row.get("km_fim")),
                _optional_float(
                    row.get("quilometragem")
                    if row.get("quilometragem") is not None
                    else row.get("km_rodado")
                ),
                _launch_status(row),
                row.get("observacao") or "",
                _humanize(row.get("tipo_entrada")),
                _safe_file_reference(row.get("caminho_arquivo")),
                _parse_datetime(row.get("recebido_em")),
            )
        )

    _prepare_sheet(sheet)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 1).number_format = DATE_FORMAT
        sheet.cell(row_index, 5).number_format = CURRENCY_FORMAT
        sheet.cell(row_index, 7).number_format = DATE_FORMAT
        sheet.cell(row_index, 12).number_format = DISTANCE_FORMAT
        sheet.cell(row_index, 13).number_format = DISTANCE_FORMAT
        sheet.cell(row_index, 14).number_format = DISTANCE_FORMAT
        sheet.cell(row_index, 19).number_format = DATETIME_FORMAT


def _build_collaborators_sheet(workbook: Workbook, rows: list[dict]) -> None:
    sheet = workbook.create_sheet("Resumo por Colaborador")
    sheet.append(
        (
            "Colaborador",
            "Total em R$",
            "Quantidade de lancamentos",
            "Quilometragem total",
            "Pendentes",
        )
    )
    for row in rows:
        sheet.append(
            (
                row.get("colaborador") or "",
                float(row.get("total") or 0),
                int(row.get("quantidade") or 0),
                float(row.get("quilometragem_total") or 0),
                int(row.get("pendentes") or 0),
            )
        )

    _prepare_sheet(sheet)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 2).number_format = CURRENCY_FORMAT
        sheet.cell(row_index, 4).number_format = DISTANCE_FORMAT


def _build_categories_sheet(workbook: Workbook, rows: list[dict]) -> None:
    sheet = workbook.create_sheet("Resumo por Categoria")
    sheet.append(("Categoria", "Total em R$", "Quantidade"))
    for row in rows:
        sheet.append(
            (
                _humanize(row.get("categoria")),
                float(row.get("total") or 0),
                int(row.get("quantidade") or 0),
            )
        )

    _prepare_sheet(sheet)
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 2).number_format = CURRENCY_FORMAT


def _prepare_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in sheet.columns:
        max_length = max(
            len(_display_value(cell.value))
            for cell in column_cells
        )
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _parse_date(value: object):
    text = str(value or "").strip()
    if not text:
        return None
    for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return text


def _parse_datetime(value: object):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value or "").strip()
    if not text:
        return None
    for date_format in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return text


def _launch_status(row: dict) -> str:
    flow_status = _humanize(row.get("status_fluxo"))
    review_status = _humanize(row.get("status_revisao"))
    return f"{flow_status} / {review_status}" if review_status else flow_status


def _safe_file_reference(value: object) -> str:
    text = str(value or "").strip()
    return Path(text).name if text else ""


def _fiscal_reference(row: dict) -> str:
    return (
        str(row.get("chave_acesso") or "").strip()
        or str(row.get("qr_code_url") or "").strip()
        or str(row.get("qr_code_text") or "").strip()
    )


def _optional_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def _humanize(value: object) -> str:
    return str(value or "").replace("_", " ").title()


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    return str(value)
